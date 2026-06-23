#!/usr/bin/env python3
"""
Script de migración: OpticaRevisado.xlsm + Cuentas.xlsx → Óptica Forever Vision API

Uso:
    python3 migrar.py --pass "Admin2026!" --url http://localhost:8001
"""

import argparse, re, sys
from datetime import datetime, date

try:
    import openpyxl, requests
except ImportError:
    import subprocess
    subprocess.run([sys.executable, "-m", "pip", "install", "openpyxl", "requests", "-q"])
    import openpyxl, requests

# ── Config ─────────────────────────────────────────────────────────────────────
parser = argparse.ArgumentParser(description="Migrar datos Excel → API")
parser.add_argument("--url",   default="http://localhost:8001")
parser.add_argument("--user",  default="admin@optica.local")
parser.add_argument("--pass",  dest="pwd", required=True)
parser.add_argument("--file",  default="OpticaRevisado.xlsm")
parser.add_argument("--file2", default="Cuentas(Recuperado automáticamente)(Recuperado automáticamente)(Recuperado automáticamente).xlsx")
parser.add_argument("--dry",   action="store_true")
args = parser.parse_args()

BASE = args.url.rstrip("/") + "/api/v1"
DRY  = args.dry

COLORS = {"g": "\033[92m", "r": "\033[91m", "y": "\033[93m", "c": "\033[96m"}
def log(msg, c=""): print(f"{COLORS.get(c,'')}{msg}\033[0m")

def parse_date(val):
    if val is None: return None
    if isinstance(val, (datetime, date)):
        return val.strftime("%Y-%m-%d") if isinstance(val, datetime) else str(val)
    s = str(val).strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y", "%d/%m/%y", "%m/%d/%Y"):
        try: return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except: pass
    return None

def clean(val):
    if val is None: return None
    s = str(val).strip()
    return None if s in ("", "None", ".", "-") else s

def safe_float(val, default=0.0):
    try: return float(str(val).replace(",", "."))
    except: return default

def split_nombre(full_name):
    parts = full_name.strip().split()
    if len(parts) <= 1: return full_name, "-"
    if len(parts) == 2: return parts[0], parts[1]
    mid = len(parts) // 2
    return " ".join(parts[:mid]), " ".join(parts[mid:])

METODOS = {"efectivo", "transferencia", "tarjeta", "cheque", "deposito", "depósito"}
def detect_metodo(*vals):
    for v in vals:
        s = str(v).strip().lower() if v else ""
        if "transfer" in s: return "transferencia"
        if "tarjet" in s:   return "tarjeta"
        if "chequ" in s:    return "cheque"
        if "depos" in s:    return "transferencia"
        if "efectivo" in s: return "efectivo"
    return "efectivo"

# ── Auth ───────────────────────────────────────────────────────────────────────
log(f"\n🔐 Autenticando en {BASE}...", "c")
resp = requests.post(f"{BASE}/auth/login", json={"email": args.user, "password": args.pwd})
if resp.status_code != 200:
    log(f"❌ Error de autenticación: {resp.text}", "r"); sys.exit(1)
token = resp.json()["access_token"]
HDR = {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}
log("✓ Autenticado", "g")

def post(endpoint, payload):
    if DRY:
        log(f"  [DRY] POST {endpoint}: {payload}", "y"); return {"id": 0}
    r = requests.post(f"{BASE}{endpoint}", json=payload, headers=HDR)
    if r.status_code not in (200, 201):
        log(f"  ⚠ {endpoint}: {r.status_code} → {r.text[:120]}", "r"); return None
    return r.json()

# ── Cuenta bancaria por defecto ────────────────────────────────────────────────
default_cuenta_id = None
r_cnt = requests.get(f"{BASE}/cuentas-bancarias", headers=HDR)
if r_cnt.status_code == 200 and r_cnt.json():
    default_cuenta_id = r_cnt.json()[0]["id"]
    log(f"  Cuenta bancaria: {r_cnt.json()[0]['nombre']} (id={default_cuenta_id})", "c")
else:
    log("❌ No hay cuentas bancarias. Crea una primero en el sistema.", "r"); sys.exit(1)

# ══════════════════════════════════════════════════════════════════════════════
# PARTE 1 — OpticaRevisado.xlsm
# ══════════════════════════════════════════════════════════════════════════════
log(f"\n📖 Leyendo {args.file}...", "c")
wb = openpyxl.load_workbook(args.file, read_only=True, data_only=True, keep_vba=False)

def sheet_data(name, header_keyword):
    ws = wb[name]
    rows = list(ws.iter_rows(values_only=True))
    hrow, hrow_idx = None, 0
    for i, row in enumerate(rows):
        if any(header_keyword.lower() in str(v).lower() for v in row if v):
            hrow = [str(v).strip() if v is not None else "" for v in row]
            hrow_idx = i; break
    if hrow is None:
        log(f"  ⚠ No se encontró '{header_keyword}' en '{name}'", "y"); return []
    return [{hrow[i]: row[i] for i in range(min(len(hrow), len(row)))}
            for row in rows[hrow_idx + 1:] if any(v is not None for v in row)]

# ── 1. Productos ───────────────────────────────────────────────────────────────
log("\n📦 [1/9] Migrando Inventario → Productos...", "c")
prod_map = {}
ok = err = skip = 0
for row in sheet_data("Inventario", "código"):
    codigo = clean(row.get("código"))
    nombre = clean(row.get("nombre del producto"))
    if not nombre: skip += 1; continue
    res = post("/productos", {
        "codigo": codigo, "nombre": nombre.title(),
        "precio_venta": safe_float(row.get("Precio")),
        "precio_costo": 0,
        "stock_actual": max(0.0, safe_float(row.get("existencias"), 0)),
        "unidad": "unidad",
    })
    if res and res.get("id"): prod_map[codigo] = res["id"]; ok += 1
    else: err += 1
log(f"  ✓ {ok} productos | {err} errores | {skip} omitidos", "g")

# ── 2. Laboratorios → Proveedores ─────────────────────────────────────────────
log("\n🔬 [2/9] Migrando Laboratorios → Proveedores...", "c")
lab_map = {}
ok = err = 0
for row in sheet_data("Registro Laboratorios", "nombre"):
    key    = clean(row.get("key laboratorio"))
    nombre = clean(row.get("nombre"))
    if not nombre: continue
    res = post("/proveedores", {
        "nombre": nombre.title(), "telefono": clean(row.get("telefono")),
        "tipo": "laboratorio", "activo": True,
    })
    if res and res.get("id"): lab_map[key] = res["id"]; ok += 1
    else: err += 1
log(f"  ✓ {ok} laboratorios | {err} errores", "g")

# ── 3. Pacientes ───────────────────────────────────────────────────────────────
log("\n👥 [3/9] Migrando bdPacientes → Pacientes...", "c")
pac_map = {}
ok = err = skip = 0
for row in sheet_data("bdPacientes", "nombres"):
    key       = clean(row.get("key"))
    full_name = clean(row.get("nombres"))
    cedula    = clean(row.get("cedula"))
    if not full_name: skip += 1; continue
    if cedula:
        cedula = re.sub(r"[^\d]", "", cedula)
        if len(cedula) < 6: cedula = None
    apellidos, nombres = split_nombre(full_name)
    res = post("/pacientes", {
        "nombres": nombres.title(), "apellidos": apellidos.title(),
        "cedula": cedula or None, "telefono": clean(row.get("telefono")),
        "fecha_nacimiento": parse_date(row.get("fecha nacimiento")),
        "observaciones": clean(row.get("observaciones")),
    })
    if res and res.get("id"):
        pac_map[key] = res["id"]; ok += 1
    else:
        # 409 → buscar por cédula
        if not DRY and cedula:
            r2 = requests.get(f"{BASE}/pacientes", headers=HDR, params={"q": cedula, "limit": 10})
            if r2.status_code == 200:
                data2 = r2.json()
                items2 = data2 if isinstance(data2, list) else data2.get("items", data2.get("results", []))
                found  = [p for p in items2 if p.get("cedula") and re.sub(r"[^\d]","",str(p["cedula"])) == cedula]
                if found:
                    pac_map[key] = found[0]["id"]; ok += 1; continue
        err += 1
log(f"  ✓ {ok} pacientes | {err} errores | {skip} omitidos", "g")

# ── 4. Ventas ──────────────────────────────────────────────────────────────────
log("\n💰 [4/9] Migrando Ventas...", "c")
venta_items = sheet_data("CuerpoVenta", "key venta")
items_por_venta: dict[str, list] = {}
for row in venta_items:
    kv = clean(row.get("key venta"))
    if kv: items_por_venta.setdefault(kv, []).append(row)

venta_map = {}
ok = err = skip = 0
for row in sheet_data("cabezaVentas", "keyVenta"):
    key_venta = clean(row.get("keyVenta"))
    if not key_venta: skip += 1; continue
    paciente_id = pac_map.get(clean(row.get("key cliente")))
    its = items_por_venta.get(key_venta, [])
    items_payload = []
    for it in its:
        kp    = clean(it.get("key producto"))
        precio = safe_float(it.get("precio"))
        cant   = safe_float(it.get("cant"), 1)
        desct  = safe_float(it.get("desct"))
        desct_pct = round((desct / (precio * cant)) * 100, 2) if precio > 0 and desct > 0 else 0.0
        items_payload.append({
            "producto_id": prod_map.get(kp), "descripcion": (kp or "Servicio").replace("-", " ").title(),
            "cantidad": cant, "precio_unitario": precio, "descuento_pct": desct_pct,
        })
    if not items_payload:
        total = safe_float(row.get("total"))
        items_payload = [{"producto_id": None, "descripcion": clean(row.get("estado")) or "Venta importada",
                          "cantidad": 1, "precio_unitario": total, "descuento_pct": 0}]
    res = post("/ventas", {
        "paciente_id": paciente_id, "fecha": parse_date(row.get("fecha")) or date.today().isoformat(),
        "descuento": 0, "notas": f"Importado desde Excel — {clean(row.get('estado')) or ''}".strip(" —"),
        "items": items_payload,
    })
    if res and res.get("id"): venta_map[key_venta] = res["id"]; ok += 1
    else: err += 1
log(f"  ✓ {ok} ventas | {err} errores | {skip} omitidos", "g")

# ── 5. Cobros de ventas (bdIngresos) ──────────────────────────────────────────
log("\n💵 [5/9] Migrando bdIngresos → Cobros de Ventas...", "c")
ok = err = skip = 0
for row in sheet_data("bdIngresos", "keyIngreso"):
    key_venta = clean(row.get("foreign key venta"))
    if not key_venta: skip += 1; continue
    venta_id = venta_map.get(key_venta)
    if not venta_id: skip += 1; continue
    monto = safe_float(row.get("monto"))
    if monto <= 0: skip += 1; continue
    metodo = (clean(row.get("metodo de pago")) or "efectivo").lower()
    metodo_norm = "transferencia" if "transfer" in metodo else "tarjeta" if "tarjet" in metodo else "cheque" if "chequ" in metodo else "efectivo"
    res = post("/cobros", {
        "venta_id": venta_id, "cuenta_bancaria_id": default_cuenta_id,
        "fecha": parse_date(row.get("fecha")) or date.today().isoformat(),
        "monto": monto, "metodo_pago": metodo_norm,
        "concepto": clean(row.get("concepto")) or "Cobro importado",
    })
    if res and res.get("id"): ok += 1
    else: err += 1
log(f"  ✓ {ok} cobros | {err} errores | {skip} omitidos", "g")
cobros_ok = ok

# ── 6. Consultas (bdConsulta) ──────────────────────────────────────────────────
log("\n🩺 [6/9] Migrando bdConsulta → Consultas...", "c")

def rx_val(v):
    if v is None: return None
    s = str(v).strip()
    if s in ("", "None", "-", "."): return None
    if s.upper() in ("N", "PLANO", "PL", "0"): return 0.0
    try: return float(s.replace(",", "."))
    except: return None

def rx_eje(v):
    if v is None: return None
    try: return int(float(str(v).replace(",", ".")))
    except: return None

ws_c = wb["bdConsulta"]
c_rows = list(ws_c.iter_rows(values_only=True))
header_idx = next((i for i, r in enumerate(c_rows[:10]) if any("key consulta" in str(v).lower() for v in r if v)), None)
ok = err = skip = 0
if header_idx is None:
    log("  ⚠ No se encontró cabecera en bdConsulta", "r")
else:
    for row in c_rows[header_idx + 1:]:
        key_c   = clean(row[2])
        key_pac = clean(row[3])
        if not key_c or not any(v is not None for v in row): skip += 1; continue
        pac_id = pac_map.get(key_pac)
        if not pac_id: skip += 1; continue
        payload = {
            "fecha":            parse_date(row[4]) or date.today().isoformat(),
            "motivo_consulta":  clean(row[5]),
            "tipo":             "consulta",
            "avsc_od":          clean(row[129]) if len(row) > 129 else None,
            "avsc_oi":          clean(row[130]) if len(row) > 130 else None,
            "avcc_ao":          clean(row[131]) if len(row) > 131 else None,
            "rx_od_esf":        rx_val(row[111] if len(row) > 111 else None),
            "rx_od_cil":        rx_val(row[112] if len(row) > 112 else None),
            "rx_od_eje":        rx_eje(row[113] if len(row) > 113 else None),
            "rx_od_add":        rx_val(row[114] if len(row) > 114 else None),
            "rx_oi_esf":        rx_val(row[115] if len(row) > 115 else None),
            "rx_oi_cil":        rx_val(row[116] if len(row) > 116 else None),
            "rx_oi_eje":        rx_eje(row[117] if len(row) > 117 else None),
            "rx_oi_add":        rx_val(row[118] if len(row) > 118 else None),
            "diagnostico":      clean(row[136]) if len(row) > 136 else None,
            "plan_tratamiento": " | ".join(filter(None, [clean(row[137]) if len(row) > 137 else None, clean(row[134]) if len(row) > 134 else None])) or None,
            "observaciones":    clean(row[138]) if len(row) > 138 else None,
            "tipo_lente":       clean(row[135]) if len(row) > 135 else None,
            "tipo_armadura":    clean(row[133]) if len(row) > 133 else None,
        }
        payload = {k: v for k, v in payload.items() if v is not None}
        res = post(f"/pacientes/{pac_id}/consultas", payload)
        if res and res.get("id"): ok += 1
        else: err += 1
log(f"  ✓ {ok} consultas | {err} errores | {skip} omitidos", "g")
consultas_ok = ok
wb.close()

# ══════════════════════════════════════════════════════════════════════════════
# PARTE 2 — Cuentas.xlsx  (INGRESOS / EGRESOS / FACTURAS)
# ══════════════════════════════════════════════════════════════════════════════
import os
if not os.path.exists(args.file2):
    log(f"\n⚠ Archivo '{args.file2}' no encontrado — omitiendo Parte 2", "y")
    wb2 = None
else:
    log(f"\n📖 Leyendo {args.file2}...", "c")
    wb2 = openpyxl.load_workbook(args.file2, read_only=True, data_only=True, keep_vba=False)

def get_sheet_rows(wb_obj, name, header_row=3):
    ws = wb_obj[name]
    rows = list(ws.iter_rows(values_only=True))
    return [list(row) for row in rows[header_row:] if any(v is not None for v in row)]

ingresos_ok = egresos_ok = cxp_ok = 0

if wb2:
    # ── 7. INGRESOS → Cobros de caja ──────────────────────────────────────────
    log("\n💵 [7/9] Migrando INGRESOS → Cobros de Caja...", "c")
    ok = err = skip = 0
    for row in get_sheet_rows(wb2, "INGRESOS"):
        fecha  = parse_date(row[0])
        nombre = clean(row[1])
        apell  = clean(row[2]) if len(row) > 2 else None
        prod   = clean(row[3]) if len(row) > 3 else None
        monto  = safe_float(row[6]) if len(row) > 6 else 0
        if not fecha or monto <= 0: skip += 1; continue
        if nombre and "anterior" in nombre.lower(): skip += 1; continue
        cliente  = " ".join(filter(None, [nombre, apell])) or "Cliente"
        concepto = f"{cliente} — {prod or 'Ingreso'}"
        res = post("/cobros", {
            "cuenta_bancaria_id": default_cuenta_id,
            "fecha": fecha, "concepto": concepto[:200], "monto": monto,
            "metodo_pago": detect_metodo(*row[7:]),
            "notas": "Importado desde Cuentas.xlsx",
        })
        if res and res.get("id"): ok += 1
        else: err += 1
    log(f"  ✓ {ok} cobros | {err} errores | {skip} omitidos", "g")
    ingresos_ok = ok

    # ── 8. EGRESOS ─────────────────────────────────────────────────────────────
    log("\n📤 [8/9] Migrando EGRESOS → Egresos...", "c")
    ok = err = skip = 0
    for row in get_sheet_rows(wb2, "EGRESOS"):
        fecha    = parse_date(row[0])
        monto    = safe_float(row[5]) if len(row) > 5 else 0
        if not fecha or monto <= 0: skip += 1; continue
        categoria = clean(row[1]) or "Otros"
        distrib   = clean(row[2]) or ""
        detalle   = clean(row[6]) if len(row) > 6 else None
        concepto  = " — ".join(filter(None, [distrib, detalle])) or categoria
        res = post("/egresos", {
            "cuenta_bancaria_id": default_cuenta_id,
            "fecha": fecha, "categoria": categoria[:100], "concepto": concepto[:200],
            "monto": monto, "metodo_pago": detect_metodo(*row[7:]),
            "notas": "Importado desde Cuentas.xlsx",
        })
        if res and res.get("id"): ok += 1
        else: err += 1
    log(f"  ✓ {ok} egresos | {err} errores | {skip} omitidos", "g")
    egresos_ok = ok

    # ── 9. FACTURAS → CxP ─────────────────────────────────────────────────────
    log("\n🧾 [9/9] Migrando FACTURAS → Cuentas por Pagar...", "c")
    ok = err = skip = 0
    for row in get_sheet_rows(wb2, "FACTURAS"):
        fecha = parse_date(row[0])
        total = safe_float(row[3]) if len(row) > 3 else 0
        if not fecha or total <= 0: skip += 1; continue
        proveedor = clean(row[2]) or clean(row[1]) or "Proveedor"
        detalle   = clean(row[4]) if len(row) > 4 else None
        ref       = clean(row[5]) if len(row) > 5 else None
        res = post("/cxp", {
            "proveedor": proveedor[:200],
            "concepto":  (detalle or f"Factura {ref or ''}").strip()[:200],
            "monto_total": total, "fecha_emision": fecha,
            "referencia": ref, "notas": "Importado desde Cuentas.xlsx",
        })
        if res and res.get("id"): ok += 1
        else: err += 1
    log(f"  ✓ {ok} facturas/CxP | {err} errores | {skip} omitidos", "g")
    cxp_ok = ok
    wb2.close()

# ── Resumen final ──────────────────────────────────────────────────────────────
log(f"""
{'='*55}
✅  MIGRACIÓN COMPLETADA
{'='*55}
  [OpticaRevisado.xlsm]
  Productos:      {len(prod_map)} creados
  Laboratorios:   {len(lab_map)} creados
  Pacientes:      {len(pac_map)} mapeados
  Ventas:         {len(venta_map)} creadas
  Cobros ventas:  {cobros_ok} creados
  Consultas:      {consultas_ok} creadas

  [Cuentas.xlsx]
  Cobros de caja: {ingresos_ok} creados
  Egresos:        {egresos_ok} creados
  CxP facturas:   {cxp_ok} creadas
{'='*55}
""", "g")
