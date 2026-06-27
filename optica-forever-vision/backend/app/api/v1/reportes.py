"""
Fase 5 — Reportes y KPIs
Endpoints:
  GET /reportes/dashboard        → KPIs generales
  GET /reportes/ventas           → detalle de ventas por rango
  GET /reportes/ventas/excel     → descarga Excel de ventas
  GET /reportes/cobros           → detalle de cobros por rango
  GET /reportes/cobros/excel     → descarga Excel de cobros
  GET /reportes/inventario       → stock actual + alertas
  GET /reportes/inventario/excel → descarga Excel de inventario
  GET /reportes/ordenes          → órdenes por estado y rango
"""
import io
from datetime import date, datetime, timedelta

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy import func, select, text
from sqlalchemy.orm import Session

from app.core.db import get_db
from app.core.deps import get_current_user
from app.models.agenda import OrdenTrabajo, Turno
from app.models.consulta import Consulta
from app.models.credito import Credito, CuotaCredito
from app.models.paciente import Paciente
from app.models.producto import Producto
from app.models.tesoreria import Cobro, Egreso
from app.models.user import User
from app.models.venta import Venta, VentaItem

router = APIRouter(prefix="/reportes", tags=["reportes"])


# ── helpers ────────────────────────────────────────────────────────────────────

def _excel_response(wb: Workbook, filename: str) -> StreamingResponse:
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    headers = {"Content-Disposition": f'attachment; filename="{filename}"'}
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


def _header_row(ws, cols: list[str], fill_hex: str = "2563EB"):
    fill = PatternFill("solid", fgColor=fill_hex)
    font = Font(bold=True, color="FFFFFF")
    for i, col in enumerate(cols, 1):
        cell = ws.cell(row=1, column=i, value=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")


# ── Dashboard KPIs ─────────────────────────────────────────────────────────────

@router.get("/dashboard")
def dashboard_kpis(
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    hoy = date.today()
    inicio_mes = hoy.replace(day=1)

    # Ventas del mes
    ventas_mes = db.execute(
        select(func.coalesce(func.sum(Venta.total), 0)).where(
            Venta.fecha >= inicio_mes,
            Venta.estado != "anulado",
        )
    ).scalar()

    # Ventas de hoy
    ventas_hoy = db.execute(
        select(func.coalesce(func.sum(Venta.total), 0)).where(
            Venta.fecha == hoy,
            Venta.estado != "anulado",
        )
    ).scalar()

    # Cobros del mes
    cobros_mes = db.execute(
        select(func.coalesce(func.sum(Cobro.monto), 0)).where(
            Cobro.fecha >= inicio_mes
        )
    ).scalar()

    # Ventas pendientes de cobro
    ventas_pendientes = db.execute(
        select(func.count()).where(Venta.estado == "pendiente")
    ).scalar()

    # Turnos de hoy
    turnos_hoy = db.execute(
        select(func.count()).where(Turno.fecha == hoy)
    ).scalar()

    # Órdenes pendientes/en proceso
    ordenes_activas = db.execute(
        select(func.count()).where(
            OrdenTrabajo.estado.in_(["pendiente", "enviado", "en_proceso"])
        )
    ).scalar()

    # Órdenes listas (para entregar)
    ordenes_listas = db.execute(
        select(func.count()).where(OrdenTrabajo.estado == "listo")
    ).scalar()

    # Pacientes nuevos este mes
    pacientes_mes = db.execute(
        select(func.count()).where(
            func.date(Paciente.created_at) >= inicio_mes
        )
    ).scalar()

    # Egresos del mes
    egresos_mes = db.execute(
        select(func.coalesce(func.sum(Egreso.monto), 0)).where(
            Egreso.fecha >= inicio_mes
        )
    ).scalar()

    # Cantidad de ventas del mes
    cant_ventas_mes = db.execute(
        select(func.count()).where(
            Venta.fecha >= inicio_mes,
            Venta.estado != "anulado",
        )
    ).scalar()

    # Cuotas vencidas
    cuotas_vencidas_count = db.execute(
        select(func.count()).where(CuotaCredito.estado == "vencido")
    ).scalar()

    cuotas_vencidas_total = db.execute(
        select(func.coalesce(
            func.sum(CuotaCredito.monto - CuotaCredito.monto_pagado), 0
        )).where(CuotaCredito.estado == "vencido")
    ).scalar()

    # Stock bajo
    stock_bajo_count = db.execute(
        select(func.count()).where(
            Producto.activo.is_(True),
            Producto.stock_actual <= Producto.stock_minimo,
        )
    ).scalar()

    # Cobros de hoy
    cobros_hoy = db.execute(
        select(func.coalesce(func.sum(Cobro.monto), 0)).where(
            Cobro.fecha == hoy
        )
    ).scalar()

    # ── Comparativo mes anterior ───────────────────────────────────────────────
    fin_mes_ant = inicio_mes - timedelta(days=1)
    inicio_mes_ant = fin_mes_ant.replace(day=1)

    ventas_mes_ant = db.execute(
        select(func.coalesce(func.sum(Venta.total), 0)).where(
            Venta.fecha.between(inicio_mes_ant, fin_mes_ant),
            Venta.estado != "anulado",
        )
    ).scalar()

    cobros_mes_ant = db.execute(
        select(func.coalesce(func.sum(Cobro.monto), 0)).where(
            Cobro.fecha.between(inicio_mes_ant, fin_mes_ant)
        )
    ).scalar()

    egresos_mes_ant = db.execute(
        select(func.coalesce(func.sum(Egreso.monto), 0)).where(
            Egreso.fecha.between(inicio_mes_ant, fin_mes_ant)
        )
    ).scalar()

    pacientes_mes_ant = db.execute(
        select(func.count()).where(
            func.date(Paciente.created_at).between(inicio_mes_ant, fin_mes_ant)
        )
    ).scalar()

    def _pct(curr: float, prev: float) -> float | None:
        if prev == 0:
            return None
        return round((curr - prev) / prev * 100, 1)

    return {
        "ventas_hoy": float(ventas_hoy),
        "ventas_mes": float(ventas_mes),
        "cobros_mes": float(cobros_mes),
        "egresos_mes": float(egresos_mes),
        "resultado_mes": float(cobros_mes) - float(egresos_mes),
        "cant_ventas_mes": int(cant_ventas_mes),
        "ventas_pendientes_cobro": int(ventas_pendientes),
        "turnos_hoy": int(turnos_hoy),
        "ordenes_activas": int(ordenes_activas),
        "ordenes_listas": int(ordenes_listas),
        "pacientes_nuevos_mes": int(pacientes_mes),
        "mes": inicio_mes.strftime("%B %Y"),
        "cuotas_vencidas_count": int(cuotas_vencidas_count),
        "cuotas_vencidas_total": float(cuotas_vencidas_total),
        "stock_bajo_count": int(stock_bajo_count),
        "cobros_hoy": float(cobros_hoy),
        # Comparativo
        "ventas_mes_ant": float(ventas_mes_ant),
        "cobros_mes_ant": float(cobros_mes_ant),
        "egresos_mes_ant": float(egresos_mes_ant),
        "pacientes_nuevos_mes_ant": int(pacientes_mes_ant),
        "ventas_pct": _pct(float(ventas_mes), float(ventas_mes_ant)),
        "cobros_pct": _pct(float(cobros_mes), float(cobros_mes_ant)),
        "egresos_pct": _pct(float(egresos_mes), float(egresos_mes_ant)),
        "pacientes_pct": _pct(int(pacientes_mes), int(pacientes_mes_ant)),
    }


# ── Alertas del día ────────────────────────────────────────────────────────────

@router.get("/alertas")
def alertas_dashboard(
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    hoy = date.today()

    # 1. Cumpleaños próximos (hoy + 7 días)
    cumpleanos_proximos = []
    for delta in range(8):
        dia = hoy + timedelta(days=delta)
        pacientes = db.execute(
            select(Paciente).where(
                Paciente.fecha_nacimiento.isnot(None),
                func.extract("month", Paciente.fecha_nacimiento) == dia.month,
                func.extract("day",   Paciente.fecha_nacimiento) == dia.day,
            )
        ).scalars().all()
        for p in pacientes:
            cumpleanos_proximos.append({
                "id": p.id,
                "nombres": p.nombres,
                "apellidos": p.apellidos,
                "telefono": p.telefono,
                "fecha_nacimiento": p.fecha_nacimiento.isoformat(),
                "dias_para": delta,
                "es_hoy": delta == 0,
            })

    # 2. Controles visuales próximos (hoy → hoy + 14 días)
    consultas_rango = db.execute(
        select(Consulta)
        .where(
            Consulta.proximo_control.isnot(None),
            Consulta.proximo_control.between(hoy, hoy + timedelta(days=14)),
        )
        .order_by(Consulta.paciente_id, Consulta.fecha.desc())
    ).scalars().all()

    seen_pac: set[int] = set()
    controles_proximos = []
    for c in consultas_rango:
        if c.paciente_id in seen_pac:
            continue
        seen_pac.add(c.paciente_id)
        p = db.get(Paciente, c.paciente_id)
        if p:
            controles_proximos.append({
                "paciente_id": p.id,
                "nombres": p.nombres,
                "apellidos": p.apellidos,
                "telefono": p.telefono,
                "proximo_control": c.proximo_control.isoformat(),
                "dias_para": (c.proximo_control - hoy).days,
            })
    controles_proximos.sort(key=lambda x: x["dias_para"])

    # 3. Turnos de hoy con detalle
    turnos_rows = db.execute(
        select(Turno, Paciente)
        .outerjoin(Paciente, Turno.paciente_id == Paciente.id)
        .where(Turno.fecha == hoy)
        .order_by(Turno.hora_inicio)
    ).all()

    turnos_hoy = [
        {
            "id": t.id,
            "paciente_id": t.paciente_id,
            "nombres": p.nombres if p else "—",
            "apellidos": p.apellidos if p else "",
            "telefono": p.telefono if p else None,
            "hora_inicio": t.hora_inicio.strftime("%H:%M"),
            "hora_fin": t.hora_fin.strftime("%H:%M") if t.hora_fin else None,
            "motivo": t.motivo,
        }
        for t, p in turnos_rows
    ]

    # 4. Cuotas pendientes que vencen en los próximos 7 días
    cuotas_rows = db.execute(
        select(CuotaCredito, Credito, Paciente)
        .join(Credito, CuotaCredito.credito_id == Credito.id)
        .outerjoin(Paciente, Credito.paciente_id == Paciente.id)
        .where(
            CuotaCredito.estado == "pendiente",
            CuotaCredito.fecha_vencimiento.between(hoy, hoy + timedelta(days=7)),
        )
        .order_by(CuotaCredito.fecha_vencimiento)
    ).all()

    cuotas_proximas = [
        {
            "paciente_id": cr.paciente_id,
            "nombres": p.nombres if p else "—",
            "apellidos": p.apellidos if p else "",
            "telefono": p.telefono if p else None,
            "credito_numero": cr.numero,
            "numero_cuota": q.numero_cuota,
            "total_cuotas": cr.numero_cuotas,
            "monto": float(q.monto - q.monto_pagado),
            "fecha_vencimiento": q.fecha_vencimiento.isoformat(),
            "dias_para": (q.fecha_vencimiento - hoy).days,
        }
        for q, cr, p in cuotas_rows
    ]

    # 5b. Productos sin stock (agotados)
    sin_stock_rows = db.execute(
        select(Producto).where(
            Producto.activo.is_(True),
            Producto.stock_actual <= 0,
        ).order_by(Producto.nombre)
    ).scalars().all()
    productos_sin_stock = [
        {"id": p.id, "nombre": p.nombre, "stock_actual": p.stock_actual}
        for p in sin_stock_rows
    ]

    # 5. Órdenes listas sin retirar (≥ 3 días en estado "listo")
    limite_dt = datetime.now() - timedelta(days=3)
    ordenes_rows = db.execute(
        select(OrdenTrabajo, Paciente)
        .outerjoin(Paciente, OrdenTrabajo.paciente_id == Paciente.id)
        .where(
            OrdenTrabajo.estado == "listo",
            OrdenTrabajo.updated_at <= limite_dt,
        )
        .order_by(OrdenTrabajo.updated_at)
    ).all()

    ordenes_sin_retirar = [
        {
            "id": o.id,
            "numero": o.numero,
            "paciente_id": o.paciente_id,
            "nombres": p.nombres if p else "—",
            "apellidos": p.apellidos if p else "",
            "telefono": p.telefono if p else None,
            "dias_esperando": (datetime.now() - o.updated_at).days,
        }
        for o, p in ordenes_rows
    ]

    return {
        "cumpleanos_proximos": cumpleanos_proximos,
        "controles_proximos": controles_proximos,
        "turnos_hoy": turnos_hoy,
        "cuotas_proximas": cuotas_proximas,
        "ordenes_sin_retirar": ordenes_sin_retirar,
        "productos_sin_stock": productos_sin_stock,
    }


# ── Reporte de Ventas ──────────────────────────────────────────────────────────

def _query_ventas(db: Session, desde: date | None, hasta: date | None):
    stmt = (
        select(
            Venta.id, Venta.numero, Venta.fecha, Venta.total,
            Venta.descuento, Venta.estado,
            Paciente.apellidos, Paciente.nombres, Paciente.cedula,
        )
        .outerjoin(Paciente, Venta.paciente_id == Paciente.id)
        .where(Venta.estado != "anulado")
        .order_by(Venta.fecha.desc(), Venta.id.desc())
    )
    if desde:
        stmt = stmt.where(Venta.fecha >= desde)
    if hasta:
        stmt = stmt.where(Venta.fecha <= hasta)
    return db.execute(stmt).all()


@router.get("/ventas")
def reporte_ventas(
    desde: date | None = None,
    hasta: date | None = None,
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    rows = _query_ventas(db, desde, hasta)
    data = [
        {
            "id": r.id,
            "numero": r.numero,
            "fecha": r.fecha.isoformat(),
            "paciente": f"{r.apellidos or ''} {r.nombres or ''}".strip() or "—",
            "cedula": r.cedula or "—",
            "total": float(r.total),
            "descuento": float(r.descuento or 0),
            "estado": r.estado,
        }
        for r in rows
    ]
    total = sum(d["total"] for d in data)
    return {"filas": data, "total": total, "cantidad": len(data)}


@router.get("/ventas/excel")
def reporte_ventas_excel(
    desde: date | None = None,
    hasta: date | None = None,
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    rows = _query_ventas(db, desde, hasta)
    wb = Workbook()
    ws = wb.active
    ws.title = "Ventas"

    cols = ["N° Venta", "Fecha", "Paciente", "Cédula", "Total", "Descuento", "Estado"]
    _header_row(ws, cols)
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 12

    for r in rows:
        paciente = f"{r.apellidos or ''} {r.nombres or ''}".strip() or "—"
        ws.append([r.numero, r.fecha, paciente, r.cedula or "—",
                   float(r.total), float(r.descuento or 0), r.estado])

    # Totales
    last = ws.max_row
    ws.append(["", "", "", "TOTAL", f'=SUM(E2:E{last})', "", ""])
    total_row = ws.max_row
    for col in ["D", "E"]:
        ws[f"{col}{total_row}"].font = Font(bold=True)

    suffix = f"_{desde}_{hasta}" if desde or hasta else ""
    return _excel_response(wb, f"ventas{suffix}.xlsx")


# ── Reporte de Cobros ──────────────────────────────────────────────────────────

def _query_cobros(db: Session, desde: date | None, hasta: date | None):
    stmt = (
        select(
            Cobro.id, Cobro.numero, Cobro.fecha, Cobro.monto,
            Cobro.metodo_pago, Cobro.referencia,
            Paciente.apellidos, Paciente.nombres,
        )
        .outerjoin(Venta, Cobro.venta_id == Venta.id)
        .outerjoin(Paciente, Venta.paciente_id == Paciente.id)
        .order_by(Cobro.fecha.desc(), Cobro.id.desc())
    )
    if desde:
        stmt = stmt.where(Cobro.fecha >= desde)
    if hasta:
        stmt = stmt.where(Cobro.fecha <= hasta)
    return db.execute(stmt).all()


@router.get("/cobros")
def reporte_cobros(
    desde: date | None = None,
    hasta: date | None = None,
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    rows = _query_cobros(db, desde, hasta)
    data = [
        {
            "id": r.id,
            "numero": r.numero,
            "fecha": r.fecha.isoformat(),
            "paciente": f"{r.apellidos or ''} {r.nombres or ''}".strip() or "—",
            "monto": float(r.monto),
            "forma_pago": r.metodo_pago,
            "referencia": r.referencia or "",
        }
        for r in rows
    ]
    total = sum(d["monto"] for d in data)
    # Subtotales por forma de pago
    por_forma: dict[str, float] = {}
    for d in data:
        por_forma[d["forma_pago"]] = por_forma.get(d["forma_pago"], 0) + d["monto"]
    return {"filas": data, "total": total, "cantidad": len(data), "por_forma_pago": por_forma}


@router.get("/cobros/excel")
def reporte_cobros_excel(
    desde: date | None = None,
    hasta: date | None = None,
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    rows = _query_cobros(db, desde, hasta)
    wb = Workbook()
    ws = wb.active
    ws.title = "Cobros"

    cols = ["N° Cobro", "Fecha", "Paciente", "Monto", "Forma de Pago", "Referencia"]
    _header_row(ws, cols)
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 20

    for r in rows:
        paciente = f"{r.apellidos or ''} {r.nombres or ''}".strip() or "—"
        ws.append([r.numero, r.fecha, paciente, float(r.monto), r.metodo_pago, r.referencia or ""])

    last = ws.max_row
    ws.append(["", "", "TOTAL", f'=SUM(D2:D{last})', "", ""])
    total_row = ws.max_row
    for col in ["C", "D"]:
        ws[f"{col}{total_row}"].font = Font(bold=True)

    suffix = f"_{desde}_{hasta}" if desde or hasta else ""
    return _excel_response(wb, f"cobros{suffix}.xlsx")


# ── Reporte de Inventario ──────────────────────────────────────────────────────

@router.get("/inventario")
def reporte_inventario(
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    rows = db.execute(
        select(Producto).where(Producto.activo == True).order_by(Producto.nombre)
    ).scalars().all()

    data = [
        {
            "id": p.id,
            "codigo": p.codigo or "",
            "nombre": p.nombre,
            "categoria_id": p.categoria_id,
            "stock_actual": p.stock_actual,
            "stock_minimo": p.stock_minimo,
            "precio_venta": float(p.precio_venta),
            "precio_costo": float(p.precio_costo) if p.precio_costo else None,
            "alerta": p.stock_actual <= p.stock_minimo,
            "valor_inventario": float(p.precio_costo or p.precio_venta) * float(p.stock_actual),
        }
        for p in rows
    ]
    valor_total = sum(d["valor_inventario"] for d in data)
    alertas = sum(1 for d in data if d["alerta"])
    return {"filas": data, "valor_total": valor_total, "total_productos": len(data), "alertas_stock": alertas}


@router.get("/inventario/excel")
def reporte_inventario_excel(
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    rows = db.execute(
        select(Producto).where(Producto.activo == True).order_by(Producto.nombre)
    ).scalars().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Inventario"

    cols = ["Código", "Nombre", "Stock actual", "Stock mínimo", "Precio venta", "Precio costo", "Valor inventario", "Alerta"]
    _header_row(ws, cols)
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 18
    ws.column_dimensions["H"].width = 10

    alert_fill = PatternFill("solid", fgColor="FEE2E2")

    for p in rows:
        alerta = p.stock_actual <= p.stock_minimo
        valor = float(p.precio_costo or p.precio_venta) * p.stock_actual
        row_data = [
            p.codigo or "", p.nombre, p.stock_actual, p.stock_minimo,
            float(p.precio_venta),
            float(p.precio_costo) if p.precio_costo else "",
            valor,
            "⚠ Bajo stock" if alerta else "",
        ]
        ws.append(row_data)
        if alerta:
            for col in range(1, 9):
                ws.cell(ws.max_row, col).fill = alert_fill

    return _excel_response(wb, "inventario.xlsx")


# ── Reporte de Órdenes ─────────────────────────────────────────────────────────

@router.get("/ordenes")
def reporte_ordenes(
    desde: date | None = None,
    hasta: date | None = None,
    estado: str | None = None,
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    stmt = (
        select(
            OrdenTrabajo.id, OrdenTrabajo.numero, OrdenTrabajo.fecha_envio,
            OrdenTrabajo.fecha_entrega_est, OrdenTrabajo.fecha_entrega_real,
            OrdenTrabajo.estado, OrdenTrabajo.tipo, OrdenTrabajo.lab_proveedor,
            OrdenTrabajo.precio_lab,
            Paciente.apellidos, Paciente.nombres,
        )
        .join(Paciente, OrdenTrabajo.paciente_id == Paciente.id)
        .order_by(OrdenTrabajo.id.desc())
    )
    if desde:
        stmt = stmt.where(OrdenTrabajo.fecha_envio >= desde)
    if hasta:
        stmt = stmt.where(OrdenTrabajo.fecha_envio <= hasta)
    if estado:
        stmt = stmt.where(OrdenTrabajo.estado == estado)

    rows = db.execute(stmt).all()
    data = [
        {
            "numero": r.numero,
            "paciente": f"{r.apellidos} {r.nombres}",
            "lab_proveedor": r.lab_proveedor,
            "tipo": r.tipo,
            "fecha_envio": r.fecha_envio.isoformat(),
            "fecha_entrega_est": r.fecha_entrega_est.isoformat() if r.fecha_entrega_est else None,
            "fecha_entrega_real": r.fecha_entrega_real.isoformat() if r.fecha_entrega_real else None,
            "estado": r.estado,
            "precio_lab": float(r.precio_lab) if r.precio_lab else None,
        }
        for r in rows
    ]
    total_lab = sum(d["precio_lab"] for d in data if d["precio_lab"])
    return {"filas": data, "total_lab": total_lab, "cantidad": len(data)}


@router.get("/proformas")
def reporte_proformas(
    desde: date | None = None,
    hasta: date | None = None,
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    stmt = (
        select(
            OrdenTrabajo.id, OrdenTrabajo.numero, OrdenTrabajo.fecha_envio,
            OrdenTrabajo.tipo, OrdenTrabajo.lab_proveedor, OrdenTrabajo.estado,
            OrdenTrabajo.precio_venta, OrdenTrabajo.precio_lab,
            OrdenTrabajo.es_proforma, OrdenTrabajo.venta_id,
            Paciente.apellidos, Paciente.nombres,
        )
        .join(Paciente, OrdenTrabajo.paciente_id == Paciente.id)
        .order_by(OrdenTrabajo.fecha_envio.desc())
    )
    if desde:
        stmt = stmt.where(OrdenTrabajo.fecha_envio >= desde)
    if hasta:
        stmt = stmt.where(OrdenTrabajo.fecha_envio <= hasta)

    rows = db.execute(stmt).all()
    filas = [
        {
            "numero": r.numero,
            "paciente": f"{r.apellidos} {r.nombres}",
            "tipo": r.tipo,
            "lab_proveedor": r.lab_proveedor,
            "estado": r.estado,
            "fecha_envio": r.fecha_envio.isoformat(),
            "precio_venta": float(r.precio_venta) if r.precio_venta else None,
            "precio_lab": float(r.precio_lab) if r.precio_lab else None,
            "es_proforma": r.es_proforma,
            "facturada": r.venta_id is not None,
        }
        for r in rows
    ]

    facturadas = [f for f in filas if f["facturada"]]
    proformas = [f for f in filas if f["es_proforma"]]
    pendientes = [f for f in filas if not f["facturada"] and not f["es_proforma"]]

    return {
        "filas": filas,
        "total_facturado": sum(f["precio_venta"] or 0 for f in facturadas),
        "total_proforma": sum(f["precio_venta"] or 0 for f in proformas),
        "total_pendiente": sum(f["precio_venta"] or 0 for f in pendientes),
        "cant_facturadas": len(facturadas),
        "cant_proformas": len(proformas),
        "cant_pendientes": len(pendientes),
    }


@router.get("/analytics")
def analytics(
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    hoy = date.today()
    result: dict = {
        "ventas_por_mes": [],
        "cobros_por_mes": [],
        "top_productos": [],
        "pacientes_por_origen": [],
        "cobros_por_metodo": [],
        "pacientes_por_mes": [],
        "ordenes_por_estado": [],
    }

    try:
        rows = db.execute(text("""
            SELECT to_char(fecha, 'YYYY-MM') AS mes,
                   SUM(total) AS total,
                   COUNT(*) AS cantidad
            FROM ventas
            WHERE estado != :estado AND fecha >= :desde
            GROUP BY to_char(fecha, 'YYYY-MM')
            ORDER BY 1
        """), {"estado": "anulado", "desde": hoy.replace(day=1) - timedelta(days=365)}).all()
        result["ventas_por_mes"] = [
            {"mes": r.mes, "total": float(r.total), "cantidad": int(r.cantidad)}
            for r in rows
        ]
    except Exception:
        db.rollback()

    try:
        rows = db.execute(text("""
            SELECT to_char(fecha, 'YYYY-MM') AS mes,
                   SUM(monto) AS total,
                   COUNT(*) AS cantidad
            FROM cobros
            WHERE fecha >= :desde
            GROUP BY to_char(fecha, 'YYYY-MM')
            ORDER BY 1
        """), {"desde": hoy.replace(day=1) - timedelta(days=365)}).all()
        result["cobros_por_mes"] = [
            {"mes": r.mes, "total": float(r.total), "cantidad": int(r.cantidad)}
            for r in rows
        ]
    except Exception:
        db.rollback()

    try:
        rows = db.execute(
            select(
                VentaItem.descripcion,
                func.sum(VentaItem.cantidad).label("cantidad"),
                func.sum(VentaItem.subtotal).label("total"),
            )
            .group_by(VentaItem.descripcion)
            .order_by(func.sum(VentaItem.subtotal).desc())
            .limit(10)
        ).all()
        result["top_productos"] = [
            {"nombre": r.descripcion, "cantidad": float(r.cantidad), "total": float(r.total)}
            for r in rows
        ]
    except Exception:
        db.rollback()

    try:
        rows = db.execute(text("""
            SELECT COALESCE(origen, 'No especificado') AS origen,
                   COUNT(*) AS cantidad
            FROM pacientes
            GROUP BY COALESCE(origen, 'No especificado')
            ORDER BY COUNT(*) DESC
        """)).all()
        result["pacientes_por_origen"] = [
            {"origen": r.origen, "cantidad": int(r.cantidad)} for r in rows
        ]
    except Exception:
        db.rollback()

    try:
        rows = db.execute(
            select(
                Cobro.metodo_pago,
                func.sum(Cobro.monto).label("total"),
                func.count().label("cantidad"),
            )
            .where(Cobro.fecha >= hoy - timedelta(days=90))
            .group_by(Cobro.metodo_pago)
            .order_by(func.sum(Cobro.monto).desc())
        ).all()
        result["cobros_por_metodo"] = [
            {"metodo": r.metodo_pago, "total": float(r.total), "cantidad": int(r.cantidad)}
            for r in rows
        ]
    except Exception:
        db.rollback()

    try:
        rows = db.execute(text("""
            SELECT to_char(created_at, 'YYYY-MM') AS mes,
                   COUNT(*) AS cantidad
            FROM pacientes
            WHERE created_at >= :desde
            GROUP BY to_char(created_at, 'YYYY-MM')
            ORDER BY 1
        """), {"desde": hoy.replace(day=1) - timedelta(days=180)}).all()
        result["pacientes_por_mes"] = [
            {"mes": r.mes, "cantidad": int(r.cantidad)} for r in rows
        ]
    except Exception:
        db.rollback()

    try:
        rows = db.execute(
            select(
                OrdenTrabajo.estado,
                func.count().label("cantidad"),
            )
            .group_by(OrdenTrabajo.estado)
        ).all()
        result["ordenes_por_estado"] = [
            {"estado": r.estado, "cantidad": int(r.cantidad)} for r in rows
        ]
    except Exception:
        db.rollback()

    return result


# ── Reporte Anual completo ─────────────────────────────────────────────────────

@router.get("/anual")
def reporte_anual(
    year: int = Query(default=None),
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    if year is None:
        year = date.today().year

    inicio = date(year, 1, 1)
    fin = date(year, 12, 31)
    meses_labels = ["Ene", "Feb", "Mar", "Abr", "May", "Jun",
                    "Jul", "Ago", "Sep", "Oct", "Nov", "Dic"]

    # ── Ventas por mes ─────────────────────────────────────────────────────────
    rows_v = db.execute(text("""
        SELECT EXTRACT(MONTH FROM fecha)::int AS mes,
               COALESCE(SUM(total), 0) AS total,
               COUNT(*) AS cantidad
        FROM ventas
        WHERE estado != 'anulado'
          AND fecha BETWEEN :inicio AND :fin
        GROUP BY EXTRACT(MONTH FROM fecha)
        ORDER BY mes
    """), {"inicio": inicio, "fin": fin}).all()
    ventas_dict = {r.mes: (float(r.total), int(r.cantidad)) for r in rows_v}

    # ── Cobros por mes ─────────────────────────────────────────────────────────
    rows_c = db.execute(text("""
        SELECT EXTRACT(MONTH FROM fecha)::int AS mes,
               COALESCE(SUM(monto), 0) AS total
        FROM cobros
        WHERE fecha BETWEEN :inicio AND :fin
        GROUP BY EXTRACT(MONTH FROM fecha)
        ORDER BY mes
    """), {"inicio": inicio, "fin": fin}).all()
    cobros_dict = {r.mes: float(r.total) for r in rows_c}

    # ── Egresos por mes ────────────────────────────────────────────────────────
    rows_e = db.execute(text("""
        SELECT EXTRACT(MONTH FROM fecha)::int AS mes,
               COALESCE(SUM(monto), 0) AS total
        FROM egresos
        WHERE fecha BETWEEN :inicio AND :fin
        GROUP BY EXTRACT(MONTH FROM fecha)
        ORDER BY mes
    """), {"inicio": inicio, "fin": fin}).all()
    egresos_dict = {r.mes: float(r.total) for r in rows_e}

    # ── Pacientes nuevos por mes ───────────────────────────────────────────────
    rows_p = db.execute(text("""
        SELECT EXTRACT(MONTH FROM created_at)::int AS mes,
               COUNT(*) AS cantidad
        FROM pacientes
        WHERE created_at BETWEEN :inicio AND :fin
        GROUP BY EXTRACT(MONTH FROM created_at)
        ORDER BY mes
    """), {"inicio": inicio, "fin": fin}).all()
    pac_dict = {r.mes: int(r.cantidad) for r in rows_p}

    # ── Consultas por mes ──────────────────────────────────────────────────────
    rows_con = db.execute(text("""
        SELECT EXTRACT(MONTH FROM fecha)::int AS mes,
               COUNT(*) AS cantidad
        FROM consultas
        WHERE fecha BETWEEN :inicio AND :fin
        GROUP BY EXTRACT(MONTH FROM fecha)
        ORDER BY mes
    """), {"inicio": inicio, "fin": fin}).all()
    cons_dict = {r.mes: int(r.cantidad) for r in rows_con}

    # ── Órdenes por mes ────────────────────────────────────────────────────────
    rows_o = db.execute(text("""
        SELECT EXTRACT(MONTH FROM fecha_envio)::int AS mes,
               COUNT(*) AS cantidad
        FROM ordenes_trabajo
        WHERE fecha_envio BETWEEN :inicio AND :fin
        GROUP BY EXTRACT(MONTH FROM fecha_envio)
        ORDER BY mes
    """), {"inicio": inicio, "fin": fin}).all()
    ord_dict = {r.mes: int(r.cantidad) for r in rows_o}

    # ── Construir series mensuales ─────────────────────────────────────────────
    por_mes = []
    for m in range(1, 13):
        v = ventas_dict.get(m, (0, 0))
        c = cobros_dict.get(m, 0)
        e = egresos_dict.get(m, 0)
        por_mes.append({
            "mes": m,
            "label": meses_labels[m - 1],
            "ventas": v[0],
            "cant_ventas": v[1],
            "cobros": c,
            "egresos": e,
            "resultado": round(c - e, 2),
            "pacientes_nuevos": pac_dict.get(m, 0),
            "consultas": cons_dict.get(m, 0),
            "ordenes": ord_dict.get(m, 0),
        })

    # ── Totales ────────────────────────────────────────────────────────────────
    total_ventas = sum(x["ventas"] for x in por_mes)
    total_cobros = sum(x["cobros"] for x in por_mes)
    total_egresos = sum(x["egresos"] for x in por_mes)
    total_pac = sum(x["pacientes_nuevos"] for x in por_mes)
    total_cons = sum(x["consultas"] for x in por_mes)
    total_ords = sum(x["ordenes"] for x in por_mes)
    total_cant_ventas = sum(x["cant_ventas"] for x in por_mes)

    # ── Top productos ──────────────────────────────────────────────────────────
    rows_tp = db.execute(text("""
        SELECT vi.descripcion,
               SUM(vi.cantidad) AS cant,
               SUM(vi.subtotal) AS total
        FROM venta_items vi
        JOIN ventas v ON vi.venta_id = v.id
        WHERE v.estado != 'anulado'
          AND v.fecha BETWEEN :inicio AND :fin
        GROUP BY vi.descripcion
        ORDER BY SUM(vi.subtotal) DESC
        LIMIT 10
    """), {"inicio": inicio, "fin": fin}).all()
    top_productos = [{"nombre": r.descripcion, "cantidad": float(r.cant), "total": float(r.total)} for r in rows_tp]

    # ── Métodos de pago ────────────────────────────────────────────────────────
    rows_mp = db.execute(text("""
        SELECT metodo_pago, SUM(monto) AS total, COUNT(*) AS cant
        FROM cobros
        WHERE fecha BETWEEN :inicio AND :fin
        GROUP BY metodo_pago
        ORDER BY SUM(monto) DESC
    """), {"inicio": inicio, "fin": fin}).all()
    metodos_pago = [{"metodo": r.metodo_pago, "total": float(r.total), "cantidad": int(r.cant)} for r in rows_mp]

    # ── Egresos por categoría ──────────────────────────────────────────────────
    rows_ec = db.execute(text("""
        SELECT categoria, SUM(monto) AS total, COUNT(*) AS cant
        FROM egresos
        WHERE fecha BETWEEN :inicio AND :fin
        GROUP BY categoria
        ORDER BY SUM(monto) DESC
    """), {"inicio": inicio, "fin": fin}).all()
    egresos_por_cat = [{"categoria": r.categoria, "total": float(r.total), "cantidad": int(r.cant)} for r in rows_ec]

    # ── Origen de pacientes ────────────────────────────────────────────────────
    rows_ori = db.execute(text("""
        SELECT COALESCE(origen, 'No especificado') AS origen, COUNT(*) AS cant
        FROM pacientes
        WHERE created_at BETWEEN :inicio AND :fin
        GROUP BY COALESCE(origen, 'No especificado')
        ORDER BY COUNT(*) DESC
    """), {"inicio": inicio, "fin": fin}).all()
    origen_pacientes = [{"origen": r.origen, "cantidad": int(r.cant)} for r in rows_ori]

    # ── Órdenes por tipo ───────────────────────────────────────────────────────
    rows_ot = db.execute(text("""
        SELECT tipo, COUNT(*) AS cant, COALESCE(SUM(precio_lab), 0) AS total_lab
        FROM ordenes_trabajo
        WHERE fecha_envio BETWEEN :inicio AND :fin
        GROUP BY tipo
        ORDER BY COUNT(*) DESC
    """), {"inicio": inicio, "fin": fin}).all()
    ordenes_por_tipo = [{"tipo": r.tipo, "cantidad": int(r.cant), "total_lab": float(r.total_lab)} for r in rows_ot]

    # ── Créditos otorgados ─────────────────────────────────────────────────────
    try:
        rows_cr = db.execute(text("""
            SELECT COUNT(*) AS cant, COALESCE(SUM(monto_total), 0) AS total
            FROM creditos
            WHERE fecha_inicio BETWEEN :inicio AND :fin
        """), {"inicio": inicio, "fin": fin}).one()
        creditos = {"cantidad": int(rows_cr.cant), "total": float(rows_cr.total)}
    except Exception:
        db.rollback()
        creditos = {"cantidad": 0, "total": 0}

    return {
        "year": year,
        "por_mes": por_mes,
        "totales": {
            "ventas": total_ventas,
            "cant_ventas": total_cant_ventas,
            "cobros": total_cobros,
            "egresos": total_egresos,
            "resultado": round(total_cobros - total_egresos, 2),
            "pacientes_nuevos": total_pac,
            "consultas": total_cons,
            "ordenes": total_ords,
            "ticket_promedio": round(total_ventas / total_cant_ventas, 2) if total_cant_ventas > 0 else 0,
        },
        "top_productos": top_productos,
        "metodos_pago": metodos_pago,
        "egresos_por_categoria": egresos_por_cat,
        "origen_pacientes": origen_pacientes,
        "ordenes_por_tipo": ordenes_por_tipo,
        "creditos": creditos,
    }


# ── Excel anual ───────────────────────────────────────────────────────────────

@router.get("/anual/excel")
def reporte_anual_excel(
    year: int = Query(default=None),
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    if year is None:
        year = date.today().year

    inicio = date(year, 1, 1)
    fin = date(year, 12, 31)
    meses_labels = ["Ene", "Feb", "Mar", "Abr", "May", "Jun",
                    "Jul", "Ago", "Sep", "Oct", "Nov", "Dic"]

    rows_v = db.execute(text("""
        SELECT EXTRACT(MONTH FROM fecha)::int AS mes, COALESCE(SUM(total), 0) AS total, COUNT(*) AS cantidad
        FROM ventas WHERE estado != 'anulado' AND fecha BETWEEN :inicio AND :fin
        GROUP BY EXTRACT(MONTH FROM fecha) ORDER BY mes
    """), {"inicio": inicio, "fin": fin}).all()
    ventas_dict = {r.mes: (float(r.total), int(r.cantidad)) for r in rows_v}

    rows_c = db.execute(text("""
        SELECT EXTRACT(MONTH FROM fecha)::int AS mes, COALESCE(SUM(monto), 0) AS total
        FROM cobros WHERE fecha BETWEEN :inicio AND :fin
        GROUP BY EXTRACT(MONTH FROM fecha) ORDER BY mes
    """), {"inicio": inicio, "fin": fin}).all()
    cobros_dict = {r.mes: float(r.total) for r in rows_c}

    rows_e = db.execute(text("""
        SELECT EXTRACT(MONTH FROM fecha)::int AS mes, COALESCE(SUM(monto), 0) AS total
        FROM egresos WHERE fecha BETWEEN :inicio AND :fin
        GROUP BY EXTRACT(MONTH FROM fecha) ORDER BY mes
    """), {"inicio": inicio, "fin": fin}).all()
    egresos_dict = {r.mes: float(r.total) for r in rows_e}

    wb = Workbook()
    ws = wb.active
    ws.title = f"Anual {year}"

    cols = ["Mes", "Ventas $", "Cant. Ventas", "Cobros $", "Egresos $", "Resultado $"]
    _header_row(ws, cols)
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 16

    pos_fill = PatternFill("solid", fgColor="D1FAE5")
    neg_fill = PatternFill("solid", fgColor="FEE2E2")

    for m in range(1, 13):
        ventas, cant = ventas_dict.get(m, (0, 0))
        cobros = cobros_dict.get(m, 0)
        egresos = egresos_dict.get(m, 0)
        resultado = cobros - egresos
        ws.append([meses_labels[m - 1], ventas, cant, cobros, egresos, resultado])
        fill = pos_fill if resultado >= 0 else neg_fill
        ws.cell(ws.max_row, 6).fill = fill

    return _excel_response(wb, f"reporte-anual-{year}.xlsx")


# ── Excel pacientes inactivos ─────────────────────────────────────────────────

@router.get("/pacientes-inactivos/excel")
def pacientes_inactivos_excel(
    meses: int = Query(default=12, ge=1, le=60),
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    from sqlalchemy import or_ as sql_or
    limite = date.today() - timedelta(days=meses * 30)

    subq = (
        select(Consulta.paciente_id, func.max(Consulta.fecha).label("ultima_consulta"))
        .group_by(Consulta.paciente_id)
        .subquery()
    )

    rows = db.execute(
        select(Paciente, subq.c.ultima_consulta)
        .outerjoin(subq, Paciente.id == subq.c.paciente_id)
        .where(
            sql_or(
                subq.c.ultima_consulta < limite,
                subq.c.ultima_consulta.is_(None),
            )
        )
        .order_by(subq.c.ultima_consulta.desc().nullslast())
        .limit(1000)
    ).all()

    hoy = date.today()
    wb = Workbook()
    ws = wb.active
    ws.title = "Pacientes Inactivos"

    cols = ["N° Paciente", "Apellidos", "Nombres", "Cédula", "Teléfono", "Última consulta", "Meses inactivo"]
    _header_row(ws, cols)
    for col, w in zip("ABCDEFG", [12, 20, 20, 14, 15, 18, 16]):
        ws.column_dimensions[col].width = w

    alert_fill = PatternFill("solid", fgColor="FEE2E2")

    for p, uc in rows:
        meses_inactivo = round((hoy - uc).days / 30) if uc else None
        row_data = [p.numero or "", p.apellidos, p.nombres, p.cedula or "", p.telefono or "", uc.isoformat() if uc else "Nunca", meses_inactivo or "Nunca"]
        ws.append(row_data)
        if meses_inactivo is None or meses_inactivo >= 12:
            for col in range(1, 8):
                ws.cell(ws.max_row, col).fill = alert_fill

    return _excel_response(wb, f"pacientes-inactivos-{meses}m.xlsx")


# ── Pacientes inactivos ────────────────────────────────────────────────────────

@router.get("/pacientes-inactivos")
def pacientes_inactivos(
    meses: int = Query(default=12, ge=1, le=60),
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    from sqlalchemy import or_ as sql_or
    limite = date.today() - timedelta(days=meses * 30)

    subq = (
        select(Consulta.paciente_id, func.max(Consulta.fecha).label("ultima_consulta"))
        .group_by(Consulta.paciente_id)
        .subquery()
    )

    rows = db.execute(
        select(Paciente, subq.c.ultima_consulta)
        .outerjoin(subq, Paciente.id == subq.c.paciente_id)
        .where(
            sql_or(
                subq.c.ultima_consulta < limite,
                subq.c.ultima_consulta.is_(None),
            )
        )
        .order_by(subq.c.ultima_consulta.desc().nullslast())
        .limit(300)
    ).all()

    hoy = date.today()
    data = [
        {
            "id": p.id,
            "numero": p.numero,
            "nombres": p.nombres,
            "apellidos": p.apellidos,
            "telefono": p.telefono,
            "ultima_consulta": uc.isoformat() if uc else None,
            "meses_inactivo": round((hoy - uc).days / 30) if uc else None,
        }
        for p, uc in rows
    ]
    return {"filas": data, "meses_filtro": meses, "total": len(data)}
