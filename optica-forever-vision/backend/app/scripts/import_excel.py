"""
Importador Fase 6 — migra datos históricos desde los Excel originales.

Archivos de origen (montados en /app/data/):
  - OpticaRevisado.xlsm → pacientes, productos, ventas, cobros
  - Cuentas.xlsx        → egresos, cuentas por pagar

Uso (desde el contenedor Docker):
  sudo docker-compose exec backend python -m app.scripts.import_excel

El script es idempotente: se puede correr varias veces sin duplicar registros.
"""
import sys
from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation

import openpyxl
from sqlalchemy import func, select, text
from sqlalchemy.orm import Session

from app.core.db import SessionLocal
from app.models.paciente import Paciente
from app.models.producto import Categoria, Producto
from app.models.tesoreria import CuentaBancaria, Cobro, CuentaPorPagar, Egreso
from app.models.user import User
from app.models.venta import Venta, VentaItem

from app.models.consulta import Consulta
from app.models.credito import Credito, CuotaCredito
OPTICA_PATH = "/app/data/OpticaRevisado.xlsm"
CUENTAS_PATH = "/app/data/Cuentas.xlsx"


# ── helpers ────────────────────────────────────────────────────────────────────

def _dec(v) -> Decimal:
    if v is None:
        return Decimal("0")
    try:
        return Decimal(str(v)).quantize(Decimal("0.01"))
    except InvalidOperation:
        return Decimal("0")


def _date(v) -> date | None:
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip()[:10]
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    return None


def _str(v, maxlen: int = 0) -> str:
    s = "" if v is None else str(v).strip()
    return s[:maxlen] if maxlen else s


def _telefono(v) -> str | None:
    if not v:
        return None
    s = str(v).strip()
    # Excel already stores "593XXXXXXXXX" → convert to local "0XXXXXXXXX"
    if s.startswith("593") and len(s) >= 12:
        s = "0" + s[3:]
    return s[:20] or None


def _split_nombre(full: str) -> tuple[str, str]:
    """'APELLIDO1 APELLIDO2 NOMBRE1 NOMBRE2' → (apellidos, nombres)."""
    parts = full.strip().split()
    if len(parts) <= 1:
        return "", full.strip()
    if len(parts) == 2:
        return parts[0], parts[1]
    if len(parts) == 3:
        return f"{parts[0]} {parts[1]}", parts[2]
    return f"{parts[0]} {parts[1]}", " ".join(parts[2:])


def _metodo_cuenta(raw: str) -> tuple[str, str]:
    """Normaliza forma de pago → (metodo_pago, nombre_cuenta_bancaria)."""
    r = (raw or "").upper().strip()
    if "EFECTIVO" in r:
        return "efectivo", "Efectivo"
    if "PICHINCHA" in r:
        return "transferencia", "Banco Pichincha"
    if "PRODUBANCO" in r:
        return "transferencia", "Produbanco"
    if "JEP" in r:
        return "transferencia", "JEP"
    if "JARDIN" in r or "JARDÍN" in r or "AZUAYO" in r:
        return "transferencia", "Jardín Azuayo"
    if "DEUNA" in r or "DE UNA" in r:
        return "transferencia", "DeUna"
    if "TARJETA" in r:
        return "tarjeta", "Banco Pichincha"
    if "TRANSFERENCIA" in r or "TRANSFER" in r:
        return "transferencia", "Banco Pichincha"
    return "efectivo", "Efectivo"


def _categoria_egreso(cuenta: str) -> str:
    c = (cuenta or "").upper()
    if any(x in c for x in ["COMIDA", "ALIMENT", "FRUTAS", "DESAY", "ALMUERZ"]):
        return "Alimentación"
    if any(x in c for x in ["LUZ", "AGUA", "INTERNET", "TELEFON"]):
        return "Luz/Agua/Internet"
    if "ARRIENDO" in c:
        return "Arriendo"
    if any(x in c for x in ["COMPRA", "INSUMO", "MATERIAL", "SUMINISTRO"]):
        return "Compras/Insumos"
    if any(x in c for x in ["BISEL", "LUNA", "LABORATOR"]):
        return "Bisel y Lunas"
    if "MOTORIZADO" in c or "MOTO" in c:
        return "Motorizado"
    if any(x in c for x in ["PERSONAL", "SUELDO", "ADELANTO", "EMPLEADO",
                              "DANNA", "BRYAM", "SOFIA", "JOSE"]):
        return "Personal/Sueldos"
    if any(x in c for x in ["PUBLICIDAD", "REDES", "MARKETING"]):
        return "Publicidad"
    if "MANTENIMIENTO" in c:
        return "Mantenimiento"
    return "Otros"


def _categoria_producto(nombre: str) -> str:
    n = nombre.upper()
    if "CONTACT" in n:
        return "Lentes de Contacto"
    if any(x in n for x in ["LUNA", "LENTE", "CR39", "POLICARB", "TRIVEX", "BLUE", "ANTIREFL", "GX7", "GX"]):
        return "Lunas"
    if any(x in n for x in ["ARMAZON", "ARMAZÓN", "MARCO", "MONTURA", "ARM."]):
        return "Armazones"
    if any(x in n for x in ["ESTUCHE", "CORDON", "PAÑO", "SOLUCION", "SOLUCIÓN", "ACCESOR"]):
        return "Accesorios"
    if any(x in n for x in ["CERTIF", "SUELDA", "SERVIC", "CONSUL", "REVISION"]):
        return "Servicios"
    return "General"


def _all_rows(ws):
    """Yield all non-blank value tuples from a worksheet."""
    for row in ws.iter_rows(values_only=True):
        if any(v is not None for v in row):
            yield row


# ── Columnas reales según diagnóstico ─────────────────────────────────────────
# bdPacientes:  [0]=None [1]=key  [2]=nombres [3]=cedula [4]=genero
#               [5]=fecha_nac [6]=edad [7]=telefono [8]=estado_civil [9]=ocupacion
#
# Inventario:   [0]=None [1]=None [2]=codigo [3]=nombre [4]=precio
#               [5]=entradas [6]=salidas [7]=existencias
#
# cabezaVentas: [0]=None [1]=None [2]=keyVenta [3]=fecha [4]=key_cliente
#               [5]=n_prods [6]=subtotal [7]=descuento [8]=total [9]=cobrado
#
# CuerpoVenta:  [0]=None [1]=None [2]=key_venta [3]=key_prod [4]=precio
#               [5]=cant [6]=descuento_$ [7]=subtotal
#
# bdIngresos:   [0]=None [1]=key_ingreso [2]=key_venta [3]=fecha
#               [4]=cliente [5]=metodo_pago [6]=monto [7]=concepto


# ── paso 1: pacientes ──────────────────────────────────────────────────────────

def import_pacientes(db: Session, ws) -> dict[str, int]:
    print("\n[1] Pacientes (bdPacientes)...")
    mapping: dict[str, int] = {}
    created = skipped = 0

    for row in _all_rows(ws):
        if len(row) < 2:
            continue
        key = _str(row[1])                          # columna B
        if not (key.startswith("P") and key[1:].isdigit()):
            continue

        num = key[1:]
        numero = f"PAC-{num.zfill(4)}"
        nombre_completo = _str(row[2])              # columna C
        cedula = _str(row[3]) or None               # columna D
        genero_raw = _str(row[4]).lower()            # columna E
        genero = genero_raw if genero_raw in {"masculino", "femenino"} else None
        fecha_nac = _date(row[5])                   # columna F (datetime o "DD/MM/YYYY")
        telefono = _telefono(row[7]) if len(row) > 7 else None   # columna H
        ocupacion = _str(row[9], 100) if len(row) > 9 else None  # columna J

        if not nombre_completo:
            continue

        existing = None
        if cedula:
            existing = db.execute(
                select(Paciente).where(Paciente.cedula == cedula)
            ).scalar_one_or_none()
        if not existing:
            existing = db.execute(
                select(Paciente).where(Paciente.numero == numero)
            ).scalar_one_or_none()

        if existing:
            mapping[key] = existing.id
            skipped += 1
            continue

        apellidos, nombres = _split_nombre(nombre_completo)
        p = Paciente(
            numero=numero,
            cedula=cedula,
            nombres=nombres or nombre_completo,
            apellidos=apellidos,
            fecha_nacimiento=fecha_nac,
            genero=genero,
            telefono=telefono,
            ocupacion=ocupacion,
        )
        db.add(p)
        db.flush()
        mapping[key] = p.id
        created += 1

    db.commit()
    print(f"  → {created} creados, {skipped} ya existían")
    return mapping


# ── paso 2: categorías ─────────────────────────────────────────────────────────

def import_categorias(db: Session) -> dict[str, int]:
    print("\n[2] Categorías de productos...")
    nombres = ["Lunas", "Armazones", "Lentes de Contacto", "Accesorios", "Servicios", "General"]
    mapping: dict[str, int] = {}
    for nombre in nombres:
        cat = db.execute(select(Categoria).where(Categoria.nombre == nombre)).scalar_one_or_none()
        if not cat:
            cat = Categoria(nombre=nombre)
            db.add(cat)
            db.flush()
            print(f"  + {nombre}")
        else:
            print(f"  = {nombre}")
        mapping[nombre] = cat.id
    db.commit()
    return mapping


# ── paso 3: productos ──────────────────────────────────────────────────────────

def import_productos(db: Session, ws, cat_map: dict[str, int]) -> dict[str, int]:
    print("\n[3] Productos (Inventario)...")
    mapping: dict[str, int] = {}
    created = skipped = 0

    HEADERS = {"código", "nombre del producto", "producto", "codigo", "nombre"}

    for row in _all_rows(ws):
        if len(row) < 4:
            continue
        codigo = _str(row[2])                       # columna C
        nombre = _str(row[3], 200)                  # columna D
        if not nombre or nombre.lower() in HEADERS:
            continue
        if not codigo or codigo.lower() in HEADERS:
            continue

        try:
            precio = _dec(row[4])                   # columna E
        except Exception:
            continue
        if precio <= 0:
            continue

        existencias_raw = row[7] if len(row) > 7 else None   # columna H
        existencias = _dec(existencias_raw)
        stock = max(existencias, Decimal("0"))

        existing = None
        if codigo:
            existing = db.execute(
                select(Producto).where(Producto.codigo == codigo)
            ).scalar_one_or_none()
        if not existing:
            existing = db.execute(
                select(Producto).where(Producto.nombre == nombre)
            ).scalar_one_or_none()

        if existing:
            mapping[codigo] = existing.id
            mapping[nombre] = existing.id
            skipped += 1
            continue

        cat_nombre = _categoria_producto(nombre)
        cat_id = cat_map.get(cat_nombre, cat_map["General"])

        p = Producto(
            codigo=codigo or None,
            nombre=nombre,
            categoria_id=cat_id,
            precio_venta=precio,
            precio_costo=Decimal("0"),
            stock_actual=stock,
            stock_minimo=Decimal("5"),
        )
        db.add(p)
        db.flush()
        mapping[codigo] = p.id
        mapping[nombre] = p.id
        created += 1

    db.commit()
    print(f"  → {created} creados, {skipped} ya existían")
    return mapping


# ── paso 4: ventas ─────────────────────────────────────────────────────────────

def import_ventas(
    db: Session,
    ws_cabeza,
    ws_cuerpo,
    pac_map: dict[str, int],
    prod_map: dict[str, int],
    admin_id: int,
) -> dict[str, int]:
    print("\n[4] Ventas + ítems (cabezaVentas + CuerpoVenta)...")

    # Pre-cargar ítems indexados por keyVenta (key_venta en columna C = índice 2)
    items_by_venta: dict[str, list] = {}
    for row in _all_rows(ws_cuerpo):
        if len(row) < 3:
            continue
        kv = _str(row[2])                           # columna C
        if kv.startswith("Venta"):
            items_by_venta.setdefault(kv, []).append(row)

    mapping: dict[str, int] = {}
    created = skipped = 0

    for row in _all_rows(ws_cabeza):
        if len(row) < 3:
            continue
        kv = _str(row[2])                           # columna C
        if not kv.startswith("Venta"):
            continue
        num = kv.replace("Venta", "")
        if not num.isdigit():
            continue
        numero = f"VEN-{num.zfill(4)}"

        existing = db.execute(
            select(Venta).where(Venta.numero == numero)
        ).scalar_one_or_none()
        if existing:
            mapping[kv] = existing.id
            skipped += 1
            continue

        fecha = _date(row[3]) or date.today()       # columna D
        pac_key = _str(row[4])                      # columna E
        subtotal = _dec(row[6]) if len(row) > 6 else Decimal("0")   # columna G
        descuento = _dec(row[7]) if len(row) > 7 else Decimal("0")  # columna H
        total = _dec(row[8]) if len(row) > 8 else Decimal("0")      # columna I
        cobrado = _dec(row[9]) if len(row) > 9 else Decimal("0")    # columna J
        estado = "cobrado" if cobrado >= total and total > 0 else "pendiente"

        venta = Venta(
            numero=numero,
            paciente_id=pac_map.get(pac_key),
            usuario_id=admin_id,
            fecha=fecha,
            subtotal=subtotal,
            descuento=descuento,
            total=total,
            estado=estado,
        )
        db.add(venta)
        db.flush()

        for irow in items_by_venta.get(kv, []):
            if len(irow) < 6:
                continue
            prod_key = _str(irow[3])                # columna D
            precio = _dec(irow[4])                  # columna E
            cant = _dec(irow[5])                    # columna F
            desct_money = _dec(irow[6]) if len(irow) > 6 else Decimal("0")  # columna G ($)
            sub = _dec(irow[7]) if len(irow) > 7 else (precio * cant - desct_money)  # columna H

            if cant <= 0 and sub <= 0:
                continue

            # Convertir descuento monetario → porcentaje
            base = precio * cant
            desct_pct = (desct_money / base * 100).quantize(Decimal("0.01")) if base > 0 and desct_money > 0 else Decimal("0")

            # Buscar nombre del producto para descripción
            prod_id = prod_map.get(prod_key)
            descripcion = prod_key[:255] if prod_key else "Producto"

            vi = VentaItem(
                venta_id=venta.id,
                producto_id=prod_id,
                descripcion=descripcion,
                cantidad=cant or Decimal("1"),
                precio_unitario=precio,
                descuento_pct=desct_pct,
                subtotal=sub,
            )
            db.add(vi)

        mapping[kv] = venta.id
        created += 1

    db.commit()
    print(f"  → {created} creadas, {skipped} ya existían")
    return mapping


# ── paso 5: cobros ─────────────────────────────────────────────────────────────

def import_cobros(
    db: Session,
    ws,
    venta_map: dict[str, int],
    cuentas: dict[str, int],
    admin_id: int,
) -> None:
    print("\n[5] Cobros (bdIngresos)...")

    existing_max = db.execute(
        text("SELECT COALESCE(MAX(CAST(SUBSTRING(numero,5) AS INTEGER)), 0) FROM cobros WHERE numero LIKE 'COB-%'")
    ).scalar()
    counter = (existing_max or 0) + 1

    created = skipped = 0

    for row in _all_rows(ws):
        if len(row) < 7:
            continue
        fecha = _date(row[3])                       # columna D (era row[2] → incorrecto)
        if not fecha:
            continue
        kv = _str(row[2])                           # columna C = key_venta (era row[1] → era keyIngreso)
        if not kv.startswith("Venta"):
            continue                                 # saltar cabeceras y filas vacías

        metodo_raw = _str(row[5]) if len(row) > 5 else "EFECTIVO"  # columna F
        monto = _dec(row[6]) if len(row) > 6 else Decimal("0")     # columna G
        concepto = _str(row[7], 255) if len(row) > 7 else ""       # columna H
        if not concepto:
            concepto = "Cobro importado"

        if monto <= 0:
            skipped += 1
            continue

        metodo, cuenta_nombre = _metodo_cuenta(metodo_raw)
        cuenta_id = cuentas.get(cuenta_nombre) or cuentas.get("Efectivo")
        venta_id = venta_map.get(kv)

        # Idempotencia: no duplicar cobros para la misma venta y monto
        if venta_id:
            ya_existe = db.execute(
                select(Cobro).where(Cobro.venta_id == venta_id, Cobro.monto == monto).limit(1)
            ).scalars().first()
            if ya_existe:
                skipped += 1
                continue

        cobro = Cobro(
            numero=f"COB-{str(counter).zfill(4)}",
            venta_id=venta_id,
            cuenta_bancaria_id=cuenta_id,
            fecha=fecha,
            concepto=concepto,
            monto=monto,
            metodo_pago=metodo,
            usuario_id=admin_id,
        )
        db.add(cobro)
        counter += 1
        created += 1

    db.execute(
        text("UPDATE configuraciones SET valor = :v WHERE clave = 'numerador_cobro'"),
        {"v": str(counter - 1)},
    )
    db.commit()
    print(f"  → {created} creados, {skipped} sin monto omitidos  (próximo: COB-{str(counter).zfill(4)})")


# ── paso 6: egresos ────────────────────────────────────────────────────────────

def import_egresos(
    db: Session,
    ws,
    cuentas: dict[str, int],
    admin_id: int,
) -> None:
    print("\n[6] Egresos (Cuentas.xlsx → EGRESOS)...")

    existing = db.execute(
        select(func.count()).select_from(Egreso).where(Egreso.numero.like("EGR-%"))
    ).scalar()
    if existing and existing > 0:
        print(f"  → {existing} egresos ya importados, saltando")
        return

    existing_max = db.execute(
        text("SELECT COALESCE(MAX(CAST(SUBSTRING(numero,5) AS INTEGER)), 0) FROM egresos WHERE numero LIKE 'EGR-%'")
    ).scalar()
    counter = (existing_max or 0) + 1
    created = skipped = 0

    for row in _all_rows(ws):
        fecha = _date(row[0]) if len(row) > 0 else None
        if not fecha:
            continue
        categoria_raw = _str(row[1]) if len(row) > 1 else ""
        if categoria_raw.upper() in {"CUENTA", ""}:
            continue
        monto = _dec(row[5]) if len(row) > 5 else Decimal("0")
        if monto <= 0:
            skipped += 1
            continue

        detalle = _str(row[6], 255) if len(row) > 6 else ""
        if not detalle:
            detalle = categoria_raw
        metodo_raw = _str(row[10]) if len(row) > 10 else "EFECTIVO"

        categoria = _categoria_egreso(categoria_raw)
        metodo, cuenta_nombre = _metodo_cuenta(metodo_raw)
        cuenta_id = cuentas.get(cuenta_nombre) or cuentas.get("Efectivo")

        egreso = Egreso(
            numero=f"EGR-{str(counter).zfill(4)}",
            cuenta_bancaria_id=cuenta_id,
            fecha=fecha,
            categoria=categoria,
            concepto=detalle[:255],
            monto=monto,
            metodo_pago=metodo,
            usuario_id=admin_id,
        )
        db.add(egreso)
        counter += 1
        created += 1

    db.execute(
        text("UPDATE configuraciones SET valor = :v WHERE clave = 'numerador_egreso'"),
        {"v": str(counter - 1)},
    )
    db.commit()
    print(f"  → {created} creados, {skipped} omitidos  (próximo: EGR-{str(counter).zfill(4)})")


# ── paso 7: cuentas por pagar ──────────────────────────────────────────────────

def import_cxp(db: Session, ws) -> None:
    print("\n[7] Cuentas por pagar (Cuentas.xlsx → FACTURAS)...")

    existing = db.execute(select(func.count()).select_from(CuentaPorPagar)).scalar()
    if existing and existing > 0:
        print(f"  → {existing} CxP ya importadas, saltando")
        return

    created = skipped = 0

    for row in _all_rows(ws):
        fecha = _date(row[0]) if len(row) > 0 else None
        if not fecha:
            continue
        estado_raw = _str(row[1]).upper()
        if estado_raw in {"CUENTA", ""}:
            continue
        proveedor = _str(row[2], 150) if len(row) > 2 else ""
        if not proveedor:
            skipped += 1
            continue
        monto = _dec(row[3]) if len(row) > 3 else Decimal("0")
        if monto <= 0:
            skipped += 1
            continue
        detalle = _str(row[4], 255) if len(row) > 4 else f"Factura {proveedor}"
        if not detalle:
            detalle = f"Factura {proveedor}"
        referencia = _str(row[5], 100) if len(row) > 5 else None

        estado = "pagado" if "PAGADO" in estado_raw else "pendiente"
        monto_pagado = monto if estado == "pagado" else Decimal("0")

        cxp = CuentaPorPagar(
            proveedor=proveedor,
            concepto=detalle,
            monto_total=monto,
            monto_pagado=monto_pagado,
            fecha_emision=fecha,
            estado=estado,
            referencia=referencia or None,
        )
        db.add(cxp)
        created += 1

    db.commit()
    print(f"  → {created} creadas, {skipped} omitidas")


# ── paso 8: laboratorios → proveedores ────────────────────────────────────────

def import_laboratorios(db: Session, ws) -> None:
    from app.models.proveedor import Proveedor
    print("\n[8] Laboratorios → Proveedores (Registro Laboratorios)...")
    created = skipped = 0

    for row in _all_rows(ws):
        if len(row) < 3:
            continue
        key    = _str(row[0])          # columna A
        nombre = _str(row[1], 150)     # columna B
        tel    = _str(row[2], 20) if len(row) > 2 else None  # columna C

        if not nombre or not key.startswith("Lab"):
            continue

        ya = db.execute(
            select(Proveedor).where(Proveedor.nombre == nombre).limit(1)
        ).scalars().first()
        if ya:
            skipped += 1
            continue

        db.add(Proveedor(nombre=nombre.title(), tipo="laboratorio", telefono=tel, activo=True))
        created += 1

    db.commit()
    print(f"  → {created} creados, {skipped} ya existían")


# ── paso 9: consultas ─────────────────────────────────────────────────────────
# bdConsulta: col C=key, D=keyPaciente, E=fecha, F=motivo
# Rx cols (0-indexed): 111=rx_od_esf, 112=rx_od_cil, 113=rx_od_eje, 114=rx_od_add
#                      115=rx_oi_esf, 116=rx_oi_cil, 117=rx_oi_eje, 118=rx_oi_add
# AV cols: 129=avsc_od, 130=avsc_oi, 131=avcc_ao
# Otros:   133=tipo_armadura, 134=plan2, 135=tipo_lente, 136=diagnostico, 137=plan1, 138=observaciones
# (índices tomados de migrar.py original)

def _rx_float(row: tuple, idx: int) -> "Decimal | None":
    if len(row) <= idx or row[idx] is None:
        return None
    s = str(row[idx]).strip()
    if s in ("", "None", "-", ".", "N", "PLANO", "PL"):
        return Decimal("0") if s in ("N", "PLANO", "PL", "0") else None
    try:
        return Decimal(s.replace(",", "."))
    except Exception:
        return None

def _rx_eje(row: tuple, idx: int) -> "int | None":
    if len(row) <= idx or row[idx] is None:
        return None
    try:
        return int(float(str(row[idx]).replace(",", ".")))
    except Exception:
        return None

def import_consultas(db: Session, ws, pac_map: dict[str, int], admin_id: int) -> None:
    print("\n[9] Consultas (bdConsulta)...")

    existing_max = db.execute(
        text("SELECT COALESCE(MAX(CAST(SUBSTRING(numero,5) AS INTEGER)), 0) FROM consultas WHERE numero LIKE 'CON-%'")
    ).scalar()
    counter = (existing_max or 0) + 1
    created = skipped = 0

    for row in _all_rows(ws):
        if len(row) < 5:
            continue
        key = _str(row[2])                        # columna C: C1, C2, ...
        if not (key.startswith("C") and key[1:].isdigit()):
            continue

        numero = f"CON-{str(counter).zfill(4)}"

        existing = db.execute(
            select(Consulta).where(Consulta.numero == numero).limit(1)
        ).scalars().first()
        if existing:
            skipped += 1
            counter += 1
            continue

        pac_key = _str(row[3])                    # columna D
        pac_id  = pac_map.get(pac_key)
        if not pac_id:
            skipped += 1
            continue

        fecha   = _date(row[4]) or date.today()   # columna E
        motivo  = _str(row[5], 500) if len(row) > 5 else None  # columna F

        # Rx data — índices exactos del migrar.py original
        plan = " | ".join(filter(None, [
            _str(row[137], 500) if len(row) > 137 else None,
            _str(row[134], 200) if len(row) > 134 else None,
        ])) or None

        c = Consulta(
            numero=numero,
            paciente_id=pac_id,
            optometrista_id=admin_id,
            fecha=fecha,
            motivo_consulta=motivo,
            avsc_od=_str(row[129], 20) if len(row) > 129 else None,
            avsc_oi=_str(row[130], 20) if len(row) > 130 else None,
            avcc_ao=_str(row[131], 20) if len(row) > 131 else None,
            rx_od_esf=_rx_float(row, 111),
            rx_od_cil=_rx_float(row, 112),
            rx_od_eje=_rx_eje(row, 113),
            rx_od_add=_rx_float(row, 114),
            rx_oi_esf=_rx_float(row, 115),
            rx_oi_cil=_rx_float(row, 116),
            rx_oi_eje=_rx_eje(row, 117),
            rx_oi_add=_rx_float(row, 118),
            diagnostico=_str(row[136], 500) if len(row) > 136 else None,
            plan_tratamiento=plan,
            observaciones=_str(row[138], 500) if len(row) > 138 else None,
            tipo_lente=_str(row[135], 100) if len(row) > 135 else None,
            tipo_armadura=_str(row[133], 100) if len(row) > 133 else None,
        )
        db.add(c)
        counter += 1
        created += 1

    db.commit()
    print(f"  → {created} creadas, {skipped} omitidas")


# ── paso 9: créditos (Cxc) ────────────────────────────────────────────────────
# Cxc: fila7=headers, fila9+ datos
# col D=key, E=fecha, F=keyCliente, G=keyVenta, H=TotalDeuda, I=Abono, J=Saldo, K=Concepto

def import_creditos(
    db: Session,
    ws,
    pac_map: dict[str, int],
    venta_map: dict[str, int],
) -> None:
    print("\n[10] Créditos (Cxc)...")

    existing_max = db.execute(
        text("SELECT COALESCE(MAX(CAST(SUBSTRING(numero,5) AS INTEGER)), 0) FROM creditos WHERE numero LIKE 'CXC-%'")
    ).scalar()
    counter = (existing_max or 0) + 1
    created = skipped = 0

    for row in _all_rows(ws):
        if len(row) < 4:
            continue
        key = _str(row[3])                         # columna D
        if not key.startswith("Cxc"):
            continue

        numero = f"CXC-{str(counter).zfill(4)}"

        existing = db.execute(
            select(Credito).where(Credito.numero == numero)
        ).scalar_one_or_none()
        if existing:
            skipped += 1
            counter += 1
            continue

        fecha      = _date(row[4]) or date.today()      # columna E
        pac_key    = _str(row[5]) if len(row) > 5 else ""  # columna F
        venta_key  = _str(row[6]) if len(row) > 6 else ""  # columna G
        monto_total = _dec(row[7]) if len(row) > 7 else Decimal("0")  # columna H
        abono      = _dec(row[8]) if len(row) > 8 else Decimal("0")   # columna I
        saldo      = _dec(row[9]) if len(row) > 9 else (monto_total - abono)  # columna J
        concepto   = _str(row[10], 255) if len(row) > 10 else "Crédito importado"  # columna K

        if monto_total <= 0:
            skipped += 1
            continue

        estado = "pagado" if saldo <= 0 else "vigente"

        cred = Credito(
            numero=numero,
            paciente_id=pac_map.get(pac_key),
            venta_id=venta_map.get(venta_key),
            monto_total=monto_total,
            abono_inicial=abono,
            monto_pagado=abono,
            numero_cuotas=1,
            fecha_inicio=fecha,
            estado=estado,
            notas=concepto or None,
        )
        db.add(cred)
        db.flush()

        # Crear la cuota por el saldo restante
        if saldo > 0:
            cuota = CuotaCredito(
                credito_id=cred.id,
                numero_cuota=1,
                monto=saldo,
                fecha_vencimiento=fecha + timedelta(days=30),
                monto_pagado=Decimal("0"),
                estado="pendiente",
            )
            db.add(cuota)

        counter += 1
        created += 1

    db.commit()
    print(f"  → {created} creados, {skipped} omitidos")


# ── paso 11: ingresos de caja (Cuentas.xlsx → INGRESOS) ──────────────────────
# Migra cobros sin venta asociada (caja diaria), igual que migrar_cuentas.py
# Columnas: FECHA(0) | NOMBRE(1) | APELLIDO(2) | PRODUCTO(3) | CANT(4) | PRECIO(5) | INGRESO(6) | DESC(7+)

def _detect_metodo(row: tuple, start: int = 7) -> str:
    for i in range(start, len(row)):
        s = str(row[i] or "").strip().lower()
        if "transfer" in s or "depos" in s:
            return "transferencia"
        if "tarjet" in s:
            return "tarjeta_debito"
        if "chequ" in s:
            return "cheque"
        if "efectivo" in s:
            return "efectivo"
    return "efectivo"

def import_cobros_caja(db: Session, ws, cuentas: dict[str, int], admin_id: int) -> None:
    print("\n[11] Cobros de caja (Cuentas.xlsx → INGRESOS)...")

    cuenta_id = (
        cuentas.get("Efectivo")
        or next(iter(cuentas.values()))
    )
    created = skipped = 0

    for row in _all_rows(ws):
        if len(row) < 7:
            continue
        fecha  = _date(row[0])
        nombre = _str(row[1], 100)
        if not fecha or not nombre:
            skipped += 1
            continue
        if "anterior" in nombre.lower():
            skipped += 1
            continue
        monto = _dec(row[6]) if len(row) > 6 else Decimal("0")
        if monto <= 0:
            skipped += 1
            continue

        apell    = _str(row[2], 100) if len(row) > 2 else None
        prod     = _str(row[3], 100) if len(row) > 3 else None
        cliente  = " ".join(filter(None, [nombre, apell])) or "Cliente"
        concepto = f"{cliente} — {prod or 'Ingreso'}"
        metodo   = _detect_metodo(row, 7)

        # Idempotencia: comparar fecha + monto + concepto
        ya = db.execute(
            select(Cobro).where(
                Cobro.fecha == fecha,
                Cobro.monto == monto,
                Cobro.concepto == concepto[:200],
                Cobro.venta_id.is_(None),
            ).limit(1)
        ).scalars().first()
        if ya:
            skipped += 1
            continue

        db.add(Cobro(
            venta_id=None,
            cuenta_bancaria_id=cuenta_id,
            fecha=fecha,
            monto=monto,
            metodo_pago=metodo,
            concepto=concepto[:200],
            creado_por=admin_id,
        ))
        created += 1

    db.commit()
    print(f"  → {created} creados, {skipped} omitidos")


# ── main ───────────────────────────────────────────────────────────────────────

def run() -> None:
    print("=" * 60)
    print("  Importador — Óptica Forever Vision")
    print("=" * 60)

    db = SessionLocal()
    try:
        admin = db.execute(select(User).where(User.role == "admin").limit(1)).scalar_one()
        print(f"\n  Usuario: {admin.email} (id={admin.id})")

        cuentas_rows = db.execute(select(CuentaBancaria)).scalars().all()
        cuentas: dict[str, int] = {c.nombre: c.id for c in cuentas_rows}
        print(f"  Cuentas: {', '.join(cuentas.keys())}")

        print(f"\n  Abriendo {OPTICA_PATH} ...")
        wb_optica = openpyxl.load_workbook(OPTICA_PATH, read_only=True, data_only=True, keep_vba=False)
        print(f"  Abriendo {CUENTAS_PATH} ...")
        wb_cuentas = openpyxl.load_workbook(CUENTAS_PATH, read_only=True, data_only=True)

        pac_map = import_pacientes(db, wb_optica["bdPacientes"])
        cat_map = import_categorias(db)
        prod_map = import_productos(db, wb_optica["Inventario"], cat_map)
        venta_map = import_ventas(
            db,
            wb_optica["cabezaVentas"],
            wb_optica["CuerpoVenta"],
            pac_map,
            prod_map,
            admin.id,
        )
        import_cobros(db, wb_optica["bdIngresos"], venta_map, cuentas, admin.id)
        import_egresos(db, wb_cuentas["EGRESOS"], cuentas, admin.id)
        import_cxp(db, wb_cuentas["FACTURAS"])
        import_laboratorios(db, wb_optica["Registro Laboratorios"])
        import_consultas(db, wb_optica["bdConsulta"], pac_map, admin.id)
        import_creditos(db, wb_optica["Cxc"], pac_map, venta_map)
        import_cobros_caja(db, wb_cuentas["INGRESOS"], cuentas, admin.id)

        wb_optica.close()
        wb_cuentas.close()

        from app.models.proveedor import Proveedor
        # Resumen final
        print("\n" + "=" * 60)
        n_pac  = db.execute(select(func.count()).select_from(Paciente)).scalar()
        n_prod = db.execute(select(func.count()).select_from(Producto)).scalar()
        n_lab  = db.execute(select(func.count()).select_from(Proveedor)).scalar()
        n_ven  = db.execute(select(func.count()).select_from(Venta)).scalar()
        n_cob  = db.execute(select(func.count()).select_from(Cobro)).scalar()
        n_egr  = db.execute(select(func.count()).select_from(Egreso)).scalar()
        n_cxp  = db.execute(select(func.count()).select_from(CuentaPorPagar)).scalar()
        n_con  = db.execute(select(func.count()).select_from(Consulta)).scalar()
        n_cred = db.execute(select(func.count()).select_from(Credito)).scalar()
        print(f"  Pacientes:        {n_pac}")
        print(f"  Productos:        {n_prod}")
        print(f"  Laboratorios:     {n_lab}")
        print(f"  Ventas:           {n_ven}")
        print(f"  Cobros (ventas):  {n_cob}")
        print(f"  Egresos:          {n_egr}")
        print(f"  Cuentas x Pagar:  {n_cxp}")
        print(f"  Consultas:        {n_con}")
        print(f"  Créditos (CxC):   {n_cred}")
        print(f"  Total cobros:     {n_cob}  (ventas + caja)")
        print("=" * 60)
        print("  [OK] Importación completada.")
        print("=" * 60)

    except Exception as e:
        db.rollback()
        print(f"\n  [ERROR] {e}", file=sys.stderr)
        import traceback
        traceback.print_exc()
        sys.exit(1)
    finally:
        db.close()


if __name__ == "__main__":
    run()
