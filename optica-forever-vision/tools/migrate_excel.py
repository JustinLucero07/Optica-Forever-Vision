"""
migrate_excel.py — Importa datos desde OpticaRevisado.xlsm al sistema.

Uso:
    pip install openpyxl requests
    python migrate_excel.py --file "C:/ruta/OpticaRevisado.xlsm" --url http://localhost --token TU_JWT

Pasos recomendados (en orden):
    1. pacientes
    2. inventario
    3. ventas
    4. creditos

Cada paso se puede correr individualmente con --solo=<paso>
"""

import sys
import argparse
import json
import datetime
import time
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERROR: instala openpyxl:  pip install openpyxl")
    sys.exit(1)

try:
    import requests
except ImportError:
    print("ERROR: instala requests:  pip install requests")
    sys.exit(1)


# ── Helpers ────────────────────────────────────────────────────────────────────

def fmtfecha(val) -> str | None:
    """Convierte datetime / string a 'YYYY-MM-DD'."""
    if val is None:
        return None
    if isinstance(val, (datetime.datetime, datetime.date)):
        return val.strftime("%Y-%m-%d")
    s = str(val).strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%m/%d/%Y", "%d-%m-%Y"):
        try:
            return datetime.datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            pass
    return None


def str_o_none(v) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def float_o_cero(v) -> float:
    try:
        return float(v or 0)
    except (TypeError, ValueError):
        return 0.0


def leer_hoja(wb, nombre: str, fila_cabecera: int):
    """Lee una hoja de Excel y devuelve lista de dicts usando la fila indicada como header."""
    ws = wb[nombre]
    headers = [ws.cell(fila_cabecera, c).value for c in range(1, ws.max_column + 1)]
    headers_limpios = [str(h).strip().lower() if h else f"col_{i}" for i, h in enumerate(headers)]
    filas = []
    for row in ws.iter_rows(min_row=fila_cabecera + 1, values_only=True):
        if all(v is None for v in row):
            continue
        filas.append(dict(zip(headers_limpios, row)))
    return filas


# ── API client ─────────────────────────────────────────────────────────────────

class API:
    def __init__(self, base_url: str, token: str):
        self.base = base_url.rstrip("/") + "/api/v1"
        self.h = {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}

    def post(self, path: str, data: dict) -> tuple[bool, dict]:
        r = requests.post(f"{self.base}{path}", json=data, headers=self.h, timeout=30)
        if r.status_code in (200, 201):
            return True, r.json()
        return False, {"error": r.text, "status": r.status_code}

    def get(self, path: str, params: dict | None = None) -> dict | list:
        r = requests.get(f"{self.base}{path}", params=params, headers=self.h, timeout=30)
        r.raise_for_status()
        return r.json()

    def login(self, usuario: str, password: str) -> str:
        r = requests.post(f"{self.base}/auth/login",
                          data={"username": usuario, "password": password},
                          timeout=10)
        r.raise_for_status()
        return r.json()["access_token"]


# ── Migración pacientes ────────────────────────────────────────────────────────

def migrar_pacientes(wb, api: API, verbose=True):
    """
    Hoja: bdPacientes
    Fila cabecera: 4  (key, nombres, cedula, genero, fecha nacimiento, ...)
    """
    print("\n── PACIENTES ──────────────────────────────────────────────")
    ws = wb["bdPacientes"]
    # Cabecera en fila 4, datos desde fila 5
    filas = list(ws.iter_rows(min_row=5, values_only=True))

    ok = err = skip = 0
    key_map: dict[str, int] = {}  # key_excel → id_sistema

    for fila in filas:
        # Cols: B=key, C=nombres, D=cedula, E=genero, F=fecha_nac, G=edad, H=telefono, I=estado_civil, J=ocupacion, K=observaciones
        key       = str_o_none(fila[1])   # columna B
        nombres   = str_o_none(fila[2])   # C
        cedula    = str_o_none(fila[3])   # D
        genero    = str_o_none(fila[4])   # E
        fecha_nac = fmtfecha(fila[5])     # F
        telefono  = str_o_none(fila[7])   # H
        ocupacion = str_o_none(fila[9])   # J
        observ    = str_o_none(fila[10])  # K

        if not key or not nombres:
            continue

        # Separar apellidos de nombres (el Excel guarda "APELLIDO NOMBRE")
        partes = nombres.strip().split()
        if len(partes) >= 4:
            apellidos = " ".join(partes[:2])
            noms = " ".join(partes[2:])
        elif len(partes) == 3:
            apellidos = partes[0]
            noms = " ".join(partes[1:])
        elif len(partes) == 2:
            apellidos = partes[0]
            noms = partes[1]
        else:
            apellidos = nombres
            noms = ""

        payload = {
            "apellidos": apellidos,
            "nombres": noms,
            "cedula": cedula,
            "genero": genero,
            "fecha_nacimiento": fecha_nac,
            "telefono": telefono,
            "ocupacion": ocupacion,
            "observaciones": observ,
        }

        exito, resp = api.post("/pacientes", payload)
        if exito:
            key_map[key] = resp["id"]
            ok += 1
            if verbose:
                print(f"  ✓ {key} → {resp['numero']}  {apellidos} {noms}")
        else:
            err += 1
            if verbose:
                print(f"  ✗ {key} {nombres}  → {resp}")

        time.sleep(0.03)  # evitar saturar el servidor

    print(f"\n  PACIENTES: {ok} importados, {err} errores, {skip} omitidos")
    return key_map


# ── Migración inventario ───────────────────────────────────────────────────────

def migrar_inventario(wb, api: API, verbose=True):
    """
    Hoja: Inventario
    Fila cabecera: 3  (código, nombre del producto, Precio, entradas, salidas, existencias)
    """
    print("\n── INVENTARIO ─────────────────────────────────────────────")
    ws = wb["Inventario"]

    ok = err = skip = 0

    for fila in ws.iter_rows(min_row=4, values_only=True):
        codigo   = str_o_none(fila[2])   # C
        nombre   = str_o_none(fila[3])   # D
        precio   = float_o_cero(fila[4]) # E
        stock    = int(float_o_cero(fila[7]) or 0)  # H = existencias

        if not nombre:
            skip += 1
            continue

        payload = {
            "codigo": codigo,
            "nombre": nombre,
            "precio_venta": precio,
            "precio_costo": None,
            "stock_actual": max(0, stock),
            "stock_minimo": 2,
            "activo": True,
            "categoria": "Importado",
        }

        exito, resp = api.post("/inventario", payload)
        if exito:
            ok += 1
            if verbose:
                print(f"  ✓ {codigo}  {nombre}  stock={stock}  precio=${precio}")
        else:
            # Si ya existe (409) no es error grave
            if "409" in str(resp.get("status", "")):
                skip += 1
            else:
                err += 1
                if verbose:
                    print(f"  ✗ {nombre} → {resp}")
        time.sleep(0.02)

    print(f"\n  INVENTARIO: {ok} importados, {err} errores, {skip} omitidos/duplicados")


# ── Migración ventas ───────────────────────────────────────────────────────────

def migrar_ventas(wb, api: API, pac_map: dict, verbose=True):
    """
    Hojas: cabezaVentas (fila 4 headers) + CuerpoVenta (fila 5 headers)
    """
    print("\n── VENTAS ─────────────────────────────────────────────────")

    # Leer cabeza de ventas
    ws_cab = wb["cabezaVentas"]
    ventas_raw = {}
    for fila in ws_cab.iter_rows(min_row=5, values_only=True):
        key_venta = str_o_none(fila[2])  # C = keyVenta
        if not key_venta:
            continue
        ventas_raw[key_venta] = {
            "fecha":    fmtfecha(fila[3]),      # D
            "pac_key":  str_o_none(fila[4]),    # E = key cliente
            "subtotal": float_o_cero(fila[6]),  # G
            "descuento": float_o_cero(fila[7]), # H
            "total":    float_o_cero(fila[8]),  # I
            "cobrado":  float_o_cero(fila[9]),  # J
            "metodo":   str_o_none(fila[11]),   # L
            "estado":   str_o_none(fila[12]),   # M
        }

    # Leer items (CuerpoVenta)
    ws_cv = wb["CuerpoVenta"]
    items_por_venta: dict[str, list] = {}
    for fila in ws_cv.iter_rows(min_row=6, values_only=True):
        key_venta  = str_o_none(fila[2])  # C = key venta
        key_prod   = str_o_none(fila[3])  # D = key producto
        precio     = float_o_cero(fila[4])# E
        cant       = int(float_o_cero(fila[5]) or 1)  # F
        descuento  = float_o_cero(fila[6])# G
        if not key_venta:
            continue
        items_por_venta.setdefault(key_venta, []).append({
            "descripcion": key_prod or "Producto importado",
            "cantidad": cant,
            "precio_unitario": precio,
            "descuento_pct": descuento,
        })

    ok = err = 0
    metodo_map = {
        "efectivo": "efectivo", "Efectivo": "efectivo",
        "transferencia": "transferencia", "Transferencia": "transferencia",
        "tarjeta": "tarjeta_debito", "Tarjeta": "tarjeta_debito",
        "cheque": "cheque", "Cheque": "cheque",
    }

    for key_v, v in ventas_raw.items():
        pac_id = pac_map.get(v["pac_key"])
        items  = items_por_venta.get(key_v, [])

        if not items:
            items = [{
                "descripcion": f"Venta {key_v}",
                "cantidad": 1,
                "precio_unitario": v["total"],
                "descuento_pct": 0,
            }]

        metodo = metodo_map.get(v["metodo"] or "", "efectivo")
        payload = {
            "paciente_id": pac_id,
            "fecha": v["fecha"] or datetime.date.today().strftime("%Y-%m-%d"),
            "descuento": v["descuento"],
            "notas": f"Importado desde Excel — {key_v}",
            "items": items,
        }

        exito, resp = api.post("/ventas", payload)
        if exito:
            ok += 1
            if verbose:
                print(f"  ✓ {key_v}  {v['fecha']}  ${v['total']}")
        else:
            err += 1
            if verbose:
                print(f"  ✗ {key_v} → {resp}")
        time.sleep(0.04)

    print(f"\n  VENTAS: {ok} importadas, {err} errores")


# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Migra datos de Excel al sistema Óptica")
    parser.add_argument("--file",   required=True, help="Ruta al archivo OpticaRevisado.xlsm")
    parser.add_argument("--url",    default="http://localhost", help="URL base del sistema (ej. http://localhost)")
    parser.add_argument("--token",  default="", help="JWT de autenticación (si está vacío se pedirá usuario/password)")
    parser.add_argument("--usuario",default="", help="Usuario admin para hacer login")
    parser.add_argument("--pass",   default="", dest="password", help="Contraseña admin")
    parser.add_argument("--solo",   default="", help="Ejecutar solo un paso: pacientes|inventario|ventas")
    parser.add_argument("--quiet",  action="store_true")
    args = parser.parse_args()

    archivo = Path(args.file)
    if not archivo.exists():
        print(f"ERROR: No se encuentra el archivo: {archivo}")
        sys.exit(1)

    print(f"📂 Leyendo {archivo.name}…")
    wb = openpyxl.load_workbook(str(archivo), read_only=True, data_only=True)
    print(f"   Hojas encontradas: {', '.join(wb.sheetnames)}")

    # Auth
    api = API(args.url, args.token)
    if not args.token:
        usuario  = args.usuario  or input("Usuario admin: ")
        password = args.password or input("Contraseña:   ")
        try:
            token = api.login(usuario, password)
            api.h["Authorization"] = f"Bearer {token}"
            print("✓ Login exitoso")
        except Exception as e:
            print(f"ERROR de login: {e}")
            sys.exit(1)

    verbose = not args.quiet
    solo    = args.solo.lower()

    pac_map: dict[str, int] = {}

    if not solo or solo == "pacientes":
        pac_map = migrar_pacientes(wb, api, verbose)

    if not solo or solo == "inventario":
        migrar_inventario(wb, api, verbose)

    if not solo or solo == "ventas":
        if not pac_map:
            print("⚠ No hay mapa de pacientes (corre primero el paso 'pacientes' o en conjunto).")
            print("  Intentando obtener pacientes del sistema…")
            try:
                pacs = api.get("/pacientes", {"limit": 2000})
                items_list = pacs.get("items", pacs) if isinstance(pacs, dict) else pacs
                # Nota: el mapa queda vacío porque no tenemos la clave original del Excel
                print(f"  Hay {len(items_list)} pacientes en el sistema.")
                print("  Las ventas se importarán sin paciente asociado (paciente_id=null).")
            except Exception:
                pass
        migrar_ventas(wb, api, pac_map, verbose)

    print("\n✅ Migración completada.")
    print("   Verificá los datos en http://localhost → Pacientes / Ventas / Inventario")


if __name__ == "__main__":
    main()
